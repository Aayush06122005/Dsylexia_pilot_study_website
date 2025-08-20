from flask import Flask, request, jsonify, session, flash, redirect, url_for, render_template, send_from_directory
from flask_cors import CORS
import mysql.connector
from datetime import datetime
from io import BytesIO, StringIO
import json
import bcrypt
import os
from werkzeug.utils import secure_filename
from authlib.integrations.flask_client import OAuth


# Load environment variables only in local dev (Railway will already have them set)
if not os.getenv("RAILWAY_ENVIRONMENT"):
    from dotenv import load_dotenv
    load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dyslexia_research_study_2025")
CORS(app)

DB_CONFIG = {
    "host": os.getenv("MYSQLHOST", "localhost"),
    "user": os.getenv("MYSQLUSER", "root"),
    "password": os.getenv("MYSQLPASSWORD", ""),
    "database": os.getenv("MYSQLDATABASE", "aviendbnew"),
    "port": int(os.getenv("MYSQLPORT", 3306))
}
def connect_db():
    """Establishes a connection to the MySQL database."""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        print("Connected to the database successfully!")
        return conn
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return None

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'wav', 'webm', 'mp3', 'ogg', 'm4a', 'jpg', 'jpeg', 'png', 'gif', 'bmp', 'tiff', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Set up OAuth
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.getenv("GOOGLE_CLIENT_ID", ""),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET", ""),
    server_metadata_url=os.getenv("GOOGLE_METADATA_URL", ""),  
    # access_token_url='https://accounts.google.com/o/oauth2/token',
    # access_token_params=None,
    # authorize_url='https://accounts.google.com/o/oauth2/auth',
    # authorize_params=None,
    # api_base_url='https://www.googleapis.com/oauth2/v1/',
    # userinfo_endpoint='https://www.googleapis.com/oauth2/v1/userinfo',
    client_kwargs={'scope': 'openid email profile'},
)
@app.route('/login/google')
def login_google():
    redirect_uri = url_for('google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/callback/google')
def google_callback():
    token = google.authorize_access_token()
    # user_info = google.get('userinfo').json()
    user_info = google.get('https://openidconnect.googleapis.com/v1/userinfo').json()
    session['user'] = user_info
    session['email'] = user_info['email']
    # Check if user exists, else create
    user_id = get_user_id(user_info['email'])
    if not user_id:
        # You may want to use user_info['name'] or user_info['given_name']
        create_user(user_info.get('name', ''), user_info['email'], '', True)
        user_id = get_user_id(user_info['email'])
    session['user_id'] = user_id
    # Check consent
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT consent_given FROM consent_data WHERE user_id = %s ORDER BY consent_date DESC LIMIT 1", (user_id,))
    consent = cursor.fetchone()
    cursor.close()
    conn.close()

    # Check demographics
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM demographics WHERE user_id = %s", (user_id,))
    demographics = cursor.fetchone()
    cursor.close()
    conn.close()

    if not demographics:
        return redirect(url_for('profile_setup'))

    # Get user type for dashboard redirect
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT user_type FROM users WHERE id = %s", (user_id,))
    user_type = cursor.fetchone()[0]
    cursor.close()
    conn.close()
    if user_type == 'parent':
        return redirect(url_for('parent_page'))
    elif user_type == 'school':
        return redirect(url_for('school_page'))
    else:
        return redirect(url_for('participant_dashboard'))

# def connect_db():
#     """Establishes a connection to the MySQL database."""
#     try:
#         conn = mysql.connector.connect(**DB_CONFIG)
#         print("Connected to the database successfully!")
#         return conn
#     except mysql.connector.Error as err:
#         print(f"Error: {err}")
#         return None

def create_user(name, email, password_hash, is_18_or_above, user_type='participant', parent_id=None, school_id=None):
    """Inserts a new user into the users table."""
    conn = connect_db()
    if conn:
        cursor = conn.cursor()
        query = """
        INSERT INTO users (name, email, password_hash, is_18_or_above, user_type, parent_id, school_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        try:
            cursor.execute(query, (name, email, password_hash, is_18_or_above, user_type, parent_id, school_id))
            conn.commit()
            print("User created successfully!")
            return True
        except Exception as err:
            print(f"Error: {err}")
            return False
        finally:
            cursor.close()
            conn.close()
    return False

def create_school(name, email, password_hash, address=None, phone=None):
    """Inserts a new school into the schools table."""
    conn = connect_db()
    if conn:
        cursor = conn.cursor()
        query = """
        INSERT INTO schools (name, email, password_hash, address, phone)
        VALUES (%s, %s, %s, %s, %s)
        """
        try:
            cursor.execute(query, (name, email, password_hash, address, phone))
            conn.commit()
            print("School created successfully!")
            return True
        except Exception as err:
            print(f"Error: {err}")
            return False
        finally:
            cursor.close()
            conn.close()
    return False

def get_school_id(email):
    """Fetches the school ID from the database based on the email."""
    try:
        conn = connect_db()
        if not conn:
            return None
        cursor = conn.cursor()
        query = "SELECT id FROM schools WHERE email = %s"
        cursor.execute(query, (email,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        return result[0] if result else None
    except Exception as e:
        print(f"Error: {e}")
        return None

def check_school_exists(email):
    """Check if a school with the given email already exists."""
    try:
        conn = connect_db()
        if not conn:
            return False
        cursor = conn.cursor()
        query = "SELECT id FROM schools WHERE email = %s"
        cursor.execute(query, (email,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        return result is not None
    except Exception as e:
        print(f"Error: {e}")
        return False

def get_user_id(email):
    """Fetches the user ID from the database based on the email."""
    try:
        conn = connect_db()
        if not conn:
            return None
        cursor = conn.cursor()
        query = "SELECT id FROM users WHERE email = %s"
        cursor.execute(query, (email,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        return result[0] if result else None
    except Exception as e:
        print(f"Error: {e}")
        return None

def get_user_name(user_id):
    """Fetches the user's name from the database based on user_id."""
    try:
        conn = connect_db()
        if not conn:
            return None
        cursor = conn.cursor()
        query = "SELECT name FROM users WHERE id = %s"
        cursor.execute(query, (user_id,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        return result[0] if result else None
    except Exception as e:
        print(f"Error: {e}")
        return None

def check_user_exists(email):
    """Check if a user with the given email already exists."""
    try:
        conn = connect_db()
        if not conn:
            return False
        cursor = conn.cursor()
        query = "SELECT id FROM users WHERE email = %s"
        cursor.execute(query, (email,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        return result is not None
    except Exception as e:
        print(f"Error: {e}")
        return False
    
@app.route('/school.html')
def school_page():
    # Check if user is logged in and is a school
    if 'school_id' not in session:
        flash('Please log in to access the school dashboard', 'error')
        return redirect(url_for('signin'))
    
    # Check if user is a school
    if session.get('user_type') != 'school':
        flash('Access denied. Only schools can access this page.', 'error')
        return redirect(url_for('landing'))
    
    return render_template('school.html')

@app.route('/school-signup')
def school_signup():
    return render_template('school_signup.html')
    
@app.route('/student-signin')
def student_signin():
    return render_template('student_signin.html')

@app.route('/api/student-login', methods=['POST'])
def student_login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'success': False, 'message': 'Username and password are required'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        # Username is stored in the email field for children
        query = "SELECT id, password_hash FROM users WHERE email = %s AND user_type = 'child'"
        cursor.execute(query, (username,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        if not result:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        user_id, stored_password_hash = result
        if bcrypt.checkpw(password.encode('utf-8'), stored_password_hash.encode('utf-8')):
            session['user_id'] = user_id
            session['user_type'] = 'child'
            return jsonify({'success': True, 'message': 'Login successful!'})
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    except Exception as e:
        print(f"Student login error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

        
@app.route('/parent')
def parent_page():
    # Check if user is logged in and is a parent
    if 'user_id' not in session:
        flash('Please log in to access the parent dashboard', 'error')
        return redirect(url_for('signin'))
    
    # Check if user is a parent
    conn = connect_db()
    if conn:
        cursor = conn.cursor()
        query = "SELECT user_type FROM users WHERE id = %s"
        cursor.execute(query, (session['user_id'],))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not result or result[0] != 'parent':
            flash('Access denied. Only parents can access this page.', 'error')
            return redirect(url_for('landing'))
    
    return render_template('parent.html')




def create_task_html(task_name, instructions, estimated_time, devices_required, example, main_content):
    safe_name = task_name.lower().replace(' ', '_')
    filename = os.path.join('templates', f"task_{safe_name}.html")

    # Load the base template from task11.html
    with open(os.path.join('templates', 'task11.html'), 'r', encoding='utf-8') as f:
        base_html = f.read()

    # Prepare the info popup and main content
    main_content = f"""
    <script>
        const TASK_NAME = "{task_name}";
    </script>
    <div id="task-info-popup" style="display:block;">
        <h2>{task_name}</h2>
        <p><strong>Instructions:</strong> {instructions}</p>
        <p><strong>Estimated Time:</strong> {estimated_time} minutes</p>
        <p><strong>Devices Required:</strong> {devices_required}</p>
        <p><strong>Example:</strong> {example}</p>
        <button onclick="document.getElementById('task-info-popup').style.display='none';document.getElementById('task-content').style.display='block';">I understand, proceed</button>
    </div>
    <div id="task-content" style="display:none;">
        {main_content}
    </div>
    """

    # Replace the main content placeholder in the base template
    final_html = base_html.replace('<!-- TASK_MAIN_CONTENT -->', info_popup)

    with open(filename, "w", encoding="utf-8") as f:
        f.write(final_html)

    # Dynamically add a Flask route for this task page
    route_path = f"/task_{safe_name}.html"
    template_name = f"task_{safe_name}.html"

    def render_task():
        return render_template(template_name)

    # if route_path not in [rule.rule for rule in app.url_map.iter_rules()]:
    #     app.add_url_rule(route_path, f"task_{safe_name}", render_task)



@app.route('/task_<safe_name>.html')
def task_dynamic(safe_name):
    template_name = f"task_{safe_name}.html"
    # Optionally, check if the template exists
    template_path = os.path.join('templates', template_name)
    if not os.path.exists(template_path):
        return "Task page not found.", 404
    return render_template(template_name)

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Handle user registration form submission."""
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '')
            is_18_or_above = request.form.get('eligibility') == 'on'
            if not name or not email or not password or not is_18_or_above:
                flash('All fields are required and eligibility must be confirmed', 'error')
                return render_template('signup.html')
            if check_user_exists(email):
                flash('User with this email already exists', 'error')
                return render_template('signup.html')
            password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            if create_user(name, email, password_hash, is_18_or_above):
                user_id = get_user_id(email)
                if user_id:
                    session['user_id'] = user_id
                    session['email'] = email
                flash('Registration successful!', 'success')
                return redirect(url_for('consent'))
            else:
                flash('Failed to create user', 'error')
                return render_template('signup.html')
        except Exception as e:
            print(f"Registration error: {e}")
            flash('Internal server error', 'error')
            return render_template('signup.html')
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login form submission."""
    if request.method == 'POST':
        try:
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '')
            if not email or not password:
                flash('Email and password are required', 'error')
                return render_template('signin.html')
            conn = connect_db()
            if not conn:
                flash('Database connection failed', 'error')
                return render_template('signin.html')
            cursor = conn.cursor()
            query = "SELECT id, password_hash FROM users WHERE email = %s"
            cursor.execute(query, (email,))
            result = cursor.fetchone()
            cursor.close()
            conn.close()
            if not result:
                flash('Invalid credentials', 'error')
                return render_template('signin.html')
            user_id, stored_password_hash = result
            if bcrypt.checkpw(password.encode('utf-8'), stored_password_hash.encode('utf-8')):
                session['user_id'] = user_id
                session['email'] = email
                flash('Login successful!', 'success')
                return redirect(url_for('participant_dashboard'))
            else:
                flash('Invalid credentials', 'error')
                return render_template('signin.html')
        except Exception as e:
            print(f"Login error: {e}")
            flash('Internal server error', 'error')
            return render_template('signin.html')
    return render_template('signin.html')

@app.route('/api/consent', methods=['POST'])
def api_consent():
    """API endpoint to store consent status and timestamp, and check digital signature."""
    user_id = request.args.get('child_id') or session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    consent_given = data.get('consent_given')
    consent_date = data.get('consent_date')
    signature = data.get('signature', '').strip()
    # Fetch user's name from DB
    user_name = get_user_name(user_id)
    if not user_name or signature.lower() != user_name.lower():
        return jsonify({'success': False, 'message': 'Digital signature must match your name (case-insensitive)'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        query = "INSERT INTO consent_data (user_id, consent_given, consent_date) VALUES (%s, %s, %s)"
        cursor.execute(query, (user_id, consent_given, consent_date))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Consent recorded'})
    except Exception as e:
        print(f"Consent error: {e}")
        return jsonify({'success': False, 'message': 'Failed to record consent'}), 500

@app.route('/api/demographics', methods=['POST'])
def api_demographics():
    """API endpoint to store profile setup/demographics data."""
    # Accept child_id as query param or in JSON
    child_id = request.args.get('child_id') or request.json.get('child_id')
    user_id = child_id or session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    dob = data.get('dob')
    gender = data.get('gender')
    native_language = data.get('native_language')
    education_level = data.get('education_level')
    dyslexia_status = data.get('dyslexia_status')
    try:
        conn = connect_db()
        cursor = conn.cursor()
        query = """
            INSERT INTO demographics (user_id, date_of_birth, gender, native_language, education_level, dyslexia_status)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (user_id, dob, gender, native_language, education_level, dyslexia_status))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Demographics recorded'})
    except Exception as e:
        print(f"Demographics error: {e}")
        return jsonify({'success': False, 'message': 'Failed to record demographics'}), 500
    
@app.route('/api/register', methods=['POST'])
def api_register():
    """API endpoint to handle user registration."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400
        
        # Extract form data
        email = data.get('email', '').strip()
        password = data.get('password', '')
        
        # Validation
        if not email or not password:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        # Check if user already exists
        if check_user_exists(email):
            return jsonify({'success': False, 'message': 'User with this email already exists'}), 409
        
        # Hash the password
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Get parent_id from session if this is a child registration
        parent_id = session.get('parent_id') if 'parent_id' in session else None
        
        # Create user
        if create_user(data.get('name', ''), email, password_hash, True, 'child' if parent_id else 'participant', parent_id, None):
            # Get user ID for session
            user_id = get_user_id(email)
            if user_id:
                session['user_id'] = user_id
                session['email'] = email
                # Clear parent_id from session after successful registration
                if 'parent_id' in session:
                    del session['parent_id']
            
            return jsonify({
                'success': True, 
                'message': 'Registration successful!',
                'user_id': user_id
            }), 201
        else:
            return jsonify({'success': False, 'message': 'Failed to create user'}), 500
            
    except Exception as e:
        print(f"Registration error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/login', methods=['POST'])
def api_login():
    """API endpoint to handle user login."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400
        
        email = data.get('email', '').strip()
        password = data.get('password', '')
        
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        
        # Check user credentials
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor()
        query = "SELECT id, password_hash FROM users WHERE email = %s"
        cursor.execute(query, (email,))
        result = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not result:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
        user_id, stored_password_hash = result
        
        # Verify password
        if bcrypt.checkpw(password.encode('utf-8'), stored_password_hash.encode('utf-8')):
            # Get user type
            conn = connect_db()
            if conn:
                cursor = conn.cursor()
                query = "SELECT user_type FROM users WHERE id = %s"
                cursor.execute(query, (user_id,))
                user_type_result = cursor.fetchone()
                cursor.close()
                conn.close()
                
                session['user_id'] = user_id
                session['email'] = email
                session['user_type'] = user_type_result[0] if user_type_result else 'participant'
                
                return jsonify({
                    'success': True,
                    'message': 'Login successful!',
                    'user_id': user_id,
                    'email': email,
                    'user_type': session['user_type']
                })
            else:
                return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
            
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    """API endpoint to handle user logout."""
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully'})

@app.route('/api/parent/register', methods=['POST'])
def parent_register():
    """API endpoint to handle parent registration."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400

        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')

        if not name or not email or not password:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400

        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        cursor = conn.cursor()

        # Check if parent exists
        cursor.execute("""
            SELECT id, password_hash FROM users 
            WHERE email = %s AND user_type = 'parent'
        """, (email,))
        result = cursor.fetchone()

        if not result:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'You are not enrolled by any school. Please contact your school.'}), 403

        user_id, password_hash_db = result

        if password_hash_db is not None:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'You are already registered. Please sign in.'}), 409

        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        # Update the parent record with password and name (in case name changed)
        cursor.execute("""
            UPDATE users SET name=%s, password_hash=%s WHERE id=%s
        """, (name, password_hash, user_id))
        conn.commit()
        cursor.close()
        conn.close()

        # Set session
        session['user_id'] = user_id
        session['email'] = email
        session['user_type'] = 'parent'

        return jsonify({'success': True, 'message': 'Parent registration successful!', 'user_id': user_id}), 201

    except Exception as e:
        print(f"Parent registration error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/parent/add-child', methods=['POST'])
def add_child():
    """API endpoint to add a child to a parent account."""
    try:
        # Check if user is logged in and is a parent
        if 'user_id' not in session or session.get('user_type') != 'parent':
            return jsonify({'success': False, 'message': 'Please log in as a parent'}), 401

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400

        username = data.get('username', '').strip()  # This will be stored in email field
        password = data.get('password', '')
        name = data.get('name', '').strip()

        # Validation
        if not username or not password or not name:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400

        # Check if username (email field) already exists
        if check_user_exists(username):
            return jsonify({'success': False, 'message': 'Username already exists'}), 409

        # Hash the password
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        # Create child user with parent_id
        if create_user(name, username, password_hash, True, 'child', session['user_id']):
            # Get child user ID
            child_id = get_user_id(username)
            # Add to parent_children table
            conn = connect_db()
            if conn:
                cursor = conn.cursor()
                try:
                    cursor.execute("INSERT INTO parent_children (parent_id, child_id) VALUES (%s, %s)", 
                                 (session['user_id'], child_id))
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    print(f"Error adding to parent_children: {e}")
                finally:
                    cursor.close()
                    conn.close()
            # Return child_id so frontend can redirect to consent
            return jsonify({
                'success': True, 
                'message': 'Child account created successfully!',
                'child_id': child_id
            }), 201
        else:
            return jsonify({'success': False, 'message': 'Failed to create child account'}), 500

    except Exception as e:
        print(f"Add child error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/parent/children', methods=['GET'])
def get_parent_children():
    """API endpoint to get all children of a parent."""
    try:
        # Check if user is logged in and is a parent
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': 'Please log in as a parent'}), 401
        
        conn = connect_db()
        if conn:
            cursor = conn.cursor(dictionary=True)
            query = """
            SELECT u.id, u.name, u.email, u.created_at, u.user_type
            FROM users u
            WHERE u.parent_id = %s AND u.user_type = 'child'
            ORDER BY u.created_at DESC
            """
            cursor.execute(query, (session['user_id'],))
            children = cursor.fetchall()
            cursor.close()
            conn.close()
            
            return jsonify({
                'success': True,
                'children': children
            })
        else:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
    except Exception as e:
        print(f"Get children error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/school/register', methods=['POST'])
def school_register():
    """API endpoint to handle school registration."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400
        
        # Extract form data
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')
        address = data.get('address', '').strip()
        phone = data.get('phone', '').strip()
        
        # Validation
        if not name or not email or not password:
            return jsonify({'success': False, 'message': 'Name, email, and password are required'}), 400
        
        # Check if school already exists
        if check_school_exists(email):
            return jsonify({'success': False, 'message': 'School with this email already exists'}), 409
        
        # Hash the password
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Create school
        if create_school(name, email, password_hash, address, phone):
            # Get school ID for session
            school_id = get_school_id(email)
            if school_id:
                session['school_id'] = school_id
                session['email'] = email
                session['user_type'] = 'school'
            
            return jsonify({
                'success': True, 
                'message': 'School registration successful!',
                'school_id': school_id
            }), 201
        else:
            return jsonify({'success': False, 'message': 'Failed to create school account'}), 500
            
    except Exception as e:
        print(f"School registration error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/school/login', methods=['POST'])
def school_login():
    """API endpoint to handle school login."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400
        
        email = data.get('email', '').strip()
        password = data.get('password', '')
        
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        
        # Check school credentials
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor()
        query = "SELECT id, password_hash FROM schools WHERE email = %s"
        cursor.execute(query, (email,))
        result = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not result:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
        school_id, stored_password_hash = result
        
        # Verify password
        if bcrypt.checkpw(password.encode('utf-8'), stored_password_hash.encode('utf-8')):
            session['school_id'] = school_id
            session['email'] = email
            session['user_type'] = 'school'
            
            return jsonify({
                'success': True,
                'message': 'School login successful!',
                'school_id': school_id,
                'email': email,
                'user_type': 'school'
            })
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
            
    except Exception as e:
        print(f"School login error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/school/add-parent', methods=['POST'])
def add_parent():
    """API endpoint to add a parent to a school account."""
    try:
        # Check if user is logged in and is a school
        if 'school_id' not in session:
            return jsonify({'success': False, 'message': 'Please log in as a school'}), 401
        
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data received'}), 400
        
        # Extract form data
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        
        # Validation
        if not name or not email:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        # Check if user already exists
        if check_user_exists(email):
            return jsonify({'success': False, 'message': 'User with this email already exists'}), 409
         
        # Create parent user with school_id
        if create_user(name, email, None, True, 'parent', None, session['school_id']):
            # Get parent user ID
            parent_id = get_user_id(email)
            
            # Add to school_parents table
            conn = connect_db()
            if conn:
                cursor = conn.cursor()
                try:
                    cursor.execute("INSERT INTO school_parents (school_id, parent_id) VALUES (%s, %s)", 
                                 (session['school_id'], parent_id))
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    print(f"Error adding to school_parents: {e}")
                finally:
                    cursor.close()
                    conn.close()
            
            return jsonify({
                'success': True, 
                'message': 'Parent account created successfully!',
                'parent_id': parent_id
            }), 201
        else:
            return jsonify({'success': False, 'message': 'Failed to create parent account'}), 500
            
    except Exception as e:
        print(f"Add parent error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/school/parents', methods=['GET'])
def get_school_parents():
    """API endpoint to get all parents of a school."""
    try:
        # Check if user is logged in and is a school
        if 'school_id' not in session:
            print(f"Session data: {dict(session)}")
            return jsonify({'success': False, 'message': 'Please log in as a school'}), 401
        
        print(f"School ID from session: {session['school_id']}")
        
        conn = connect_db()
        if conn:
            cursor = conn.cursor(dictionary=True)
            query = """
            SELECT u.id, u.name, u.email, u.created_at, u.user_type
            FROM users u
            WHERE u.school_id = %s AND u.user_type = 'parent'
            ORDER BY u.created_at DESC
            """
            cursor.execute(query, (session['school_id'],))
            parents = cursor.fetchall()
            cursor.close()
            conn.close()
            
            print(f"Found {len(parents)} parents for school {session['school_id']}")
            
            return jsonify({
                'success': True,
                'parents': parents
            })
        else:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
    except Exception as e:
        print(f"Get parents error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/api/user-info', methods=['GET'])
def get_user_info():
    """API endpoint to get current user information."""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        
        cursor = conn.cursor(dictionary=True)
        query = "SELECT id, email, date_of_birth, gender FROM users WHERE id = %s"
        cursor.execute(query, (user_id,))
        user = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if user:
            # Convert date to string format
            if user['date_of_birth']:
                user['date_of_birth'] = user['date_of_birth'].strftime('%Y-%m-%d')
            
            return jsonify({'success': True, 'user': user})
        else:
            return jsonify({'success': False, 'message': 'User not found'}), 404
            
    except Exception as e:
        print(f"Get user info error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error'}), 500

@app.route('/test-db')
def test_db():
    """Test database connection."""
    try:
        conn = connect_db()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
            cursor.close()
            conn.close()
            return jsonify({'success': True, 'message': 'Database connection successful!'})
        else:
            return jsonify({'success': False, 'message': 'Database connection failed'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'})

@app.route('/test-tasks')
def test_tasks():
    """Test tasks table."""
    try:
        conn = connect_db()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM tasks")
            tasks = cursor.fetchall()
            cursor.close()
            conn.close()
            return jsonify({'success': True, 'message': f'Found {len(tasks)} tasks', 'tasks': tasks})
        else:
            return jsonify({'success': False, 'message': 'Database connection failed'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'})

@app.route('/test-session')
def test_session():
    """Test session data."""
    session_data = {
        'user_id': session.get('user_id'),
        'school_id': session.get('school_id'),
        'email': session.get('email'),
        'user_type': session.get('user_type')
    }
    return jsonify(session_data)

# Serve HTML pages
@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/signup')
def signup():
    # Check if this is a child signup from parent dashboard
    parent_id = request.args.get('parent_id')
    if parent_id:
        # Verify parent exists and is logged in
        if 'user_id' not in session or str(session['user_id']) != str(parent_id):
            flash('Access denied. Please log in as the parent.', 'error')
            return redirect(url_for('signin'))
        
        # Store parent_id in session for child registration
        session['parent_id'] = int(parent_id)
    
    return render_template('signup.html')

@app.route('/parent-signup')
def parent_signup():
    return render_template('parent_signup.html')

@app.route('/signin')
def signin():
    return render_template('signin.html')

@app.route('/consent')
def consent():
    return render_template('consent.html')

@app.route('/profile-setup')
def profile_setup():
    return render_template('profile-setup.html')



@app.route('/participant-dashboard')
def participant_dashboard():
    user_id = session.get('user_id')
    user_name = None
    tasks = []
    total_tasks = 0
    progress = 0
    if user_id:
        user_name = get_user_name(user_id)
        try:
            conn = connect_db()
            cursor = conn.cursor(dictionary=True)
            
            # Define the core tasks for progress calculation
            core_tasks = [
                'Reading Aloud Task 1',
                'Typing Task', 
                'Reading Comprehension',
                'Mathematical Comprehension',
                'Writing Task',
                'Aptitude Test'
            ]
            
            # Get all available tasks from database
            cursor.execute("SELECT task_name FROM tasks")
            db_tasks = [row['task_name'] for row in cursor.fetchall()]
            
            # Get user task statuses
            cursor.execute("SELECT task_name, status FROM user_tasks WHERE user_id = %s", (user_id,))
            user_task_statuses = {t['task_name']: t['status'] for t in cursor.fetchall()}
            
            # Create task list with statuses
            for task_name in db_tasks:
                status = user_task_statuses.get(task_name, 'Not Started')
                tasks.append({'task_name': task_name, 'status': status})
            
            # Calculate progress based on core tasks only
            core_task_statuses = []
            for core_task in core_tasks:
                status = user_task_statuses.get(core_task, 'Not Started')
                core_task_statuses.append(status)
            
            total_tasks = len(core_tasks)
            completed = sum(1 for status in core_task_statuses if status == 'Completed')
            progress = int((completed / total_tasks) * 100) if total_tasks > 0 else 0
            
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Dashboard fetch tasks error: {e}")
    return render_template('participant-dashboard.html', user_name=user_name, tasks=tasks, progress=progress, total_tasks=total_tasks)



@app.route('/learn_more')
def learn_more():
    return render_template('learn_more.html')

@app.route('/extra')
def extra():
    return render_template('extra.html')

@app.route('/extra2')
def extra2():
    return render_template('extra2.html')

@app.route('/admin')
def admin_portal():
    if not session.get('is_admin'):
        flash('Please log in as admin to access this page', 'error')
        return redirect(url_for('admin_login'))
    return render_template('admin.html')

# --- Admin task categories pages ---
@app.route('/admin/tasks/category/<string:category_slug>')
def admin_tasks_by_category(category_slug: str):
    if not session.get('is_admin'):
        flash('Please log in as admin to access this page', 'error')
        return redirect(url_for('admin_login'))

    category_map = {
        'reading-aloud': {
            'title': 'Reading Aloud',
            'table': 'reading_tasks'
        },
        'typing': {
            'title': 'Typing Task',
            'table': 'typing_tasks'
        },
        'writing': {
            'title': 'Writing Task',
            'table': 'writing_tasks'
        },
        'reading-comprehension': {
            'title': 'Reading Comprehension',
            'table': 'reading_comprehension_tasks'
        },
        'mathematical-comprehension': {
            'title': 'Mathematical Comprehension',
            'table': 'mathematical_comprehension_tasks'
        },
        'aptitude': {
            'title': 'Aptitude Test',
            'table': 'aptitude_tasks'
        }
    }

    cfg = category_map.get(category_slug)
    if not cfg:
        flash('Unknown task category', 'error')
        return redirect(url_for('admin_portal'))

    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)

        # Build base query per category
        base_table = cfg['table']

        cursor.execute(f"""
            SELECT *
            FROM {base_table}
            ORDER BY age_min, id
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        # Group by age ranges (age_min, age_max)
        groups = {}
        for row in rows:
            age_min = row.get('age_min')
            age_max = row.get('age_max')
            key = (age_min, age_max)
            if key not in groups:
                label = f"Ages {age_min}+" if age_max is not None and int(age_max) >= 99 else f"Ages {age_min}-{age_max}"
                groups[key] = {
                    'label': label,
                    'age_min': age_min,
                    'tasks': []
                }
            # Common fields
            task_common = {
                'id': row.get('id'),
                'task_name': row.get('task_name'),
                'difficulty_level': row.get('difficulty_level') or '-',
                'estimated_time': row.get('estimated_time') or '-',
            }
            # Add category-specific content fields for display
            if category_slug == 'reading-aloud':
                task_common['content'] = row.get('content')
                task_common['instructions'] = row.get('instructions')
            elif category_slug == 'typing':
                task_common['prompt'] = row.get('prompt')
                task_common['instructions'] = row.get('instructions')
            elif category_slug == 'writing':
                task_common['prompt'] = row.get('prompt')
                task_common['instructions'] = row.get('instructions')
            elif category_slug == 'reading-comprehension':
                task_common['passage'] = row.get('passage')
                task_common['question1'] = row.get('question1')
                task_common['question2'] = row.get('question2')
                task_common['question3'] = row.get('question3')
                # Parse JSON options if needed
                import json as _json
                a1 = row.get('answer1_options')
                a2 = row.get('answer2_options')
                try:
                    if isinstance(a1, str):
                        a1 = _json.loads(a1)
                except Exception:
                    pass
                try:
                    if isinstance(a2, str):
                        a2 = _json.loads(a2)
                except Exception:
                    pass
                task_common['answer1_options'] = a1
                task_common['answer2_options'] = a2
                task_common['answer3_type'] = row.get('answer3_type')
                task_common['instructions'] = row.get('instructions')
            elif category_slug == 'mathematical-comprehension':
                task_common['problem_text'] = row.get('problem_text')
                task_common['question1'] = row.get('question1')
                task_common['question2'] = row.get('question2')
                task_common['question3'] = row.get('question3')
                import json as _json
                a1 = row.get('answer1_options')
                a2 = row.get('answer2_options')
                try:
                    if isinstance(a1, str):
                        a1 = _json.loads(a1)
                except Exception:
                    pass
                try:
                    if isinstance(a2, str):
                        a2 = _json.loads(a2)
                except Exception:
                    pass
                task_common['answer1_options'] = a1
                task_common['answer2_options'] = a2
                task_common['answer3_type'] = row.get('answer3_type')
                task_common['instructions'] = row.get('instructions')
            elif category_slug == 'aptitude':
                import json as _json
                task_common['instructions'] = row.get('instructions')
                task_common['estimated_time'] = row.get('estimated_time')
                # Do not expose example in admin list view
                # Two possible schemas: per-question columns OR JSON arrays per section
                if 'logical_question1' in row:
                    # No example/content exposure for aptitude
                    task_common['logical_question1'] = row.get('logical_question1')
                    task_common['logical_question2'] = row.get('logical_question2')
                    task_common['numerical_question1'] = row.get('numerical_question1')
                    task_common['numerical_question2'] = row.get('numerical_question2')
                    task_common['verbal_question1'] = row.get('verbal_question1')
                    task_common['verbal_question2'] = row.get('verbal_question2')
                    task_common['spatial_question1'] = row.get('spatial_question1')
                    task_common['spatial_question2'] = row.get('spatial_question2')
                    for field in [
                        'logical_question1_options','logical_question2_options',
                        'numerical_question1_options','numerical_question2_options',
                        'verbal_question1_options','verbal_question2_options',
                        'spatial_question1_options','spatial_question2_options'
                    ]:
                        options = row.get(field)
                        try:
                            if isinstance(options, str):
                                options = _json.loads(options)
                        except Exception:
                            options = []
                        task_common[field] = options
                else:
                    # JSON-based schema
                    def parse_section(json_text):
                        try:
                            if isinstance(json_text, str):
                                return _json.loads(json_text) or []
                            return json_text or []
                        except Exception:
                            return []
                    lr = parse_section(row.get('logical_reasoning_questions'))
                    na = parse_section(row.get('numerical_ability_questions'))
                    va = parse_section(row.get('verbal_ability_questions'))
                    sr = parse_section(row.get('spatial_reasoning_questions'))
                    def set_pair(prefix, arr):
                        q1 = arr[0] if len(arr) > 0 else {}
                        q2 = arr[1] if len(arr) > 1 else {}
                        task_common[f'{prefix}_question1'] = q1.get('question')
                        task_common[f'{prefix}_question2'] = q2.get('question')
                        task_common[f'{prefix}_question1_options'] = q1.get('options') or []
                        task_common[f'{prefix}_question2_options'] = q2.get('options') or []
                    set_pair('logical', lr)
                    set_pair('numerical', na)
                    set_pair('verbal', va)
                    set_pair('spatial', sr)
            groups[key]['tasks'].append(task_common)

        # Sort groups by age_min
        age_groups = [groups[k] for k in sorted(groups.keys(), key=lambda t: (t[0], t[1]))]

        return render_template(
            'admin_tasks_category.html',
            category_title=cfg['title'],
            category_slug=category_slug,
            age_groups=age_groups
        )
    except Exception as e:
        print(f"Admin tasks by category error: {e}")
        flash('Failed to load tasks for this category', 'error')
        return redirect(url_for('admin_portal'))


# --- Admin category CRUD APIs ---
def _category_config(category_slug: str):
    return {
        'reading-aloud': {'table': 'reading_tasks'},
        'typing': {'table': 'typing_tasks'},
        'writing': {'table': 'writing_tasks'},
        'reading-comprehension': {'table': 'reading_comprehension_tasks'},
        'mathematical-comprehension': {'table': 'mathematical_comprehension_tasks'},
        'aptitude': {'table': 'aptitude_tasks'}
    }.get(category_slug)


@app.route('/api/admin/categories/<string:category_slug>/tasks', methods=['GET', 'POST'])
def admin_category_tasks(category_slug: str):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    cfg = _category_config(category_slug)
    if not cfg:
        return jsonify({'success': False, 'message': 'Unknown category'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        table = cfg['table']
        if request.method == 'GET':
            cursor.execute(f"SELECT * FROM {table}")
            return jsonify({'success': True, 'tasks': cursor.fetchall()})
        else:
            data = request.get_json() or {}
            if category_slug == 'reading-aloud':
                cursor.execute(
                    """
                    INSERT INTO reading_tasks (task_name, age_min, age_max, difficulty_level, content, instructions, estimated_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                        data.get('content'), data.get('instructions'), data.get('estimated_time')
                    )
                )
            elif category_slug == 'typing':
                cursor.execute(
                    """
                    INSERT INTO typing_tasks (task_name, age_min, age_max, difficulty_level, prompt, instructions, word_limit_min, word_limit_max, estimated_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                        data.get('prompt'), data.get('instructions'), data.get('word_limit_min'), data.get('word_limit_max'), data.get('estimated_time')
                    )
                )
            elif category_slug == 'writing':
                cursor.execute(
                    """
                    INSERT INTO writing_tasks (task_name, age_min, age_max, difficulty_level, prompt, instructions, word_limit_min, word_limit_max, estimated_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                        data.get('prompt'), data.get('instructions'), data.get('word_limit_min'), data.get('word_limit_max'), data.get('estimated_time')
                    )
                )
            elif category_slug == 'reading-comprehension':
                import json as _json
                cursor.execute(
                    """
                    INSERT INTO reading_comprehension_tasks (task_name, age_min, age_max, difficulty_level, passage, question1, question2, question3, answer1_options, answer2_options, answer3_type, instructions, estimated_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                        data.get('passage'), data.get('question1'), data.get('question2'), data.get('question3'),
                        _json.dumps(data.get('answer1_options') or []), _json.dumps(data.get('answer2_options') or []),
                        data.get('answer3_type'), data.get('instructions'), data.get('estimated_time')
                    )
                )
            elif category_slug == 'mathematical-comprehension':
                import json as _json
                cursor.execute(
                    """
                    INSERT INTO mathematical_comprehension_tasks (task_name, age_min, age_max, difficulty_level, problem_text, question1, question2, question3, answer1_options, answer2_options, answer3_type, instructions, estimated_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                        data.get('problem_text'), data.get('question1'), data.get('question2'), data.get('question3'),
                        _json.dumps(data.get('answer1_options') or []), _json.dumps(data.get('answer2_options') or []),
                        data.get('answer3_type'), data.get('instructions'), data.get('estimated_time')
                    )
                )
            elif category_slug == 'aptitude':
                import json as _json
                # Detect schema and insert accordingly
                cursor2 = conn.cursor(dictionary=True)
                cursor2.execute("SHOW COLUMNS FROM aptitude_tasks")
                existing_columns = {row['Field'] for row in cursor2.fetchall()}
                cursor2.close()
                if 'logical_reasoning_questions' in existing_columns:
                    def section_from(prefix):
                        q1 = data.get(f'{prefix}_question1')
                        q2 = data.get(f'{prefix}_question2')
                        o1 = data.get(f'{prefix}_question1_options') or []
                        o2 = data.get(f'{prefix}_question2_options') or []
                        sec = []
                        if q1:
                            sec.append({'question': q1, 'options': o1})
                        if q2:
                            sec.append({'question': q2, 'options': o2})
                        return _json.dumps(sec)
                    cursor.execute(
                        """
                        INSERT INTO aptitude_tasks (
                            task_name, age_min, age_max, difficulty_level,
                            logical_reasoning_questions, numerical_ability_questions, verbal_ability_questions, spatial_reasoning_questions,
                            instructions, estimated_time
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                            section_from('logical'), section_from('numerical'), section_from('verbal'), section_from('spatial'),
                            data.get('instructions'), data.get('estimated_time')
                        )
                    )
                else:
                    cursor.execute(
                    """
                    INSERT INTO aptitude_tasks (
                        task_name, age_min, age_max, difficulty_level, instructions, estimated_time, example,
                        logical_question1, logical_question1_options, logical_question2, logical_question2_options,
                        numerical_question1, numerical_question1_options, numerical_question2, numerical_question2_options,
                        verbal_question1, verbal_question1_options, verbal_question2, verbal_question2_options,
                        spatial_question1, spatial_question1_options, spatial_question2, spatial_question2_options
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        data.get('task_name'), data.get('age_min'), data.get('age_max'), data.get('difficulty_level'),
                        data.get('instructions'), data.get('estimated_time'), data.get('example'),
                        data.get('logical_question1'), _json.dumps(data.get('logical_question1_options') or []),
                        data.get('logical_question2'), _json.dumps(data.get('logical_question2_options') or []),
                        data.get('numerical_question1'), _json.dumps(data.get('numerical_question1_options') or []),
                        data.get('numerical_question2'), _json.dumps(data.get('numerical_question2_options') or []),
                        data.get('verbal_question1'), _json.dumps(data.get('verbal_question1_options') or []),
                        data.get('verbal_question2'), _json.dumps(data.get('verbal_question2_options') or []),
                        data.get('spatial_question1'), _json.dumps(data.get('spatial_question1_options') or []),
                        data.get('spatial_question2'), _json.dumps(data.get('spatial_question2_options') or [])
                    )
                )
            conn.commit()
            return jsonify({'success': True})
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        print(f"Admin category POST error: {e}")
        return jsonify({'success': False, 'message': 'Failed to save task'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()


@app.route('/api/admin/categories/<string:category_slug>/tasks/<int:task_id>', methods=['GET', 'PUT', 'DELETE'])
def admin_category_task_detail(category_slug: str, task_id: int):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    cfg = _category_config(category_slug)
    if not cfg:
        return jsonify({'success': False, 'message': 'Unknown category'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        table = cfg['table']
        if request.method == 'DELETE':
            cursor.execute(f"DELETE FROM {table} WHERE id = %s", (task_id,))
            conn.commit()
            return jsonify({'success': True})
        elif request.method == 'GET':
            # Single task fetch, with schema normalization for aptitude
            getc = conn.cursor(dictionary=True)
            getc.execute(f"SELECT * FROM {table} WHERE id = %s", (task_id,))
            row = getc.fetchone()
            getc.close()
            if not row:
                return jsonify({'success': False, 'message': 'Task not found'}), 404
            if category_slug == 'aptitude':
                import json as _json
                normalized = dict(row)
                if 'logical_question1' not in row:
                    # JSON-based schema; fan-out to the fields admin editor expects
                    def parse_section(json_text):
                        try:
                            if isinstance(json_text, str):
                                return _json.loads(json_text) or []
                            return json_text or []
                        except Exception:
                            return []
                    lr = parse_section(row.get('logical_reasoning_questions'))
                    na = parse_section(row.get('numerical_ability_questions'))
                    va = parse_section(row.get('verbal_ability_questions'))
                    sr = parse_section(row.get('spatial_reasoning_questions'))
                    def add_pair(prefix, arr):
                        q1 = arr[0] if len(arr) > 0 else {}
                        q2 = arr[1] if len(arr) > 1 else {}
                        normalized[f'{prefix}_question1'] = q1.get('question')
                        normalized[f'{prefix}_question2'] = q2.get('question')
                        normalized[f'{prefix}_question1_options'] = q1.get('options') or []
                        normalized[f'{prefix}_question2_options'] = q2.get('options') or []
                    add_pair('logical', lr)
                    add_pair('numerical', na)
                    add_pair('verbal', va)
                    add_pair('spatial', sr)
                return jsonify({'success': True, 'task': normalized})
        else:
            # Partial update: fetch existing row first and merge
            data = request.get_json() or {}
            curd = conn.cursor(dictionary=True)
            curd.execute(f"SELECT * FROM {table} WHERE id = %s", (task_id,))
            existing = curd.fetchone() or {}
            curd.close()
            if category_slug == 'reading-aloud':
                cursor.execute(
                    """
                    UPDATE reading_tasks SET task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s, content=%s, instructions=%s, estimated_time=%s WHERE id=%s
                    """,
                    (
                        existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'),
                        data.get('content', existing.get('content')), data.get('instructions', existing.get('instructions')), data.get('estimated_time', existing.get('estimated_time')), task_id
                    )
                )
            elif category_slug == 'typing':
                cursor.execute(
                    """
                    UPDATE typing_tasks SET task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s, prompt=%s, instructions=%s, word_limit_min=%s, word_limit_max=%s, estimated_time=%s WHERE id=%s
                    """,
                    (
                        existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'),
                        data.get('prompt', existing.get('prompt')), data.get('instructions', existing.get('instructions')), data.get('word_limit_min', existing.get('word_limit_min')), data.get('word_limit_max', existing.get('word_limit_max')), data.get('estimated_time', existing.get('estimated_time')), task_id
                    )
                )
            elif category_slug == 'writing':
                cursor.execute(
                    """
                    UPDATE writing_tasks SET task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s, prompt=%s, instructions=%s, word_limit_min=%s, word_limit_max=%s, estimated_time=%s WHERE id=%s
                    """,
                    (
                        existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'),
                        data.get('prompt', existing.get('prompt')), data.get('instructions', existing.get('instructions')), data.get('word_limit_min', existing.get('word_limit_min')), data.get('word_limit_max', existing.get('word_limit_max')), data.get('estimated_time', existing.get('estimated_time')), task_id
                    )
                )
            elif category_slug == 'reading-comprehension':
                import json as _json
                cursor.execute(
                    """
                    UPDATE reading_comprehension_tasks SET task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s, passage=%s, question1=%s, question2=%s, question3=%s, answer1_options=%s, answer2_options=%s, answer3_type=%s, instructions=%s, estimated_time=%s WHERE id=%s
                    """,
                    (
                        existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'), data.get('passage', existing.get('passage')),
                        data.get('question1', existing.get('question1')), data.get('question2', existing.get('question2')), data.get('question3', existing.get('question3')), _json.dumps(data.get('answer1_options', existing.get('answer1_options') or [])), _json.dumps(data.get('answer2_options', existing.get('answer2_options') or [])),
                        existing.get('answer3_type'), data.get('instructions', existing.get('instructions')), data.get('estimated_time', existing.get('estimated_time')), task_id
                    )
                )
            elif category_slug == 'mathematical-comprehension':
                import json as _json
                cursor.execute(
                    """
                    UPDATE mathematical_comprehension_tasks SET task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s, problem_text=%s, question1=%s, question2=%s, question3=%s, answer1_options=%s, answer2_options=%s, answer3_type=%s, instructions=%s, estimated_time=%s WHERE id=%s
                    """,
                    (
                        existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'), data.get('problem_text', existing.get('problem_text')),
                        data.get('question1', existing.get('question1')), data.get('question2', existing.get('question2')), data.get('question3', existing.get('question3')), _json.dumps(data.get('answer1_options', existing.get('answer1_options') or [])), _json.dumps(data.get('answer2_options', existing.get('answer2_options') or [])),
                        existing.get('answer3_type'), data.get('instructions', existing.get('instructions')), data.get('estimated_time', existing.get('estimated_time')), task_id
                    )
                )
            elif category_slug == 'aptitude':
                import json as _json
                # Detect schema and update accordingly
                chk = conn.cursor(dictionary=True)
                chk.execute("SHOW COLUMNS FROM aptitude_tasks")
                existing_columns = {row['Field'] for row in chk.fetchall()}
                chk.close()
                if 'logical_reasoning_questions' in existing_columns:
                    # Build section arrays from incoming or existing values
                    def section_from(prefix):
                        # Prefer incoming discrete fields, else try to reconstruct from existing JSON
                        q1 = data.get(f'{prefix}_question1')
                        q2 = data.get(f'{prefix}_question2')
                        o1 = data.get(f'{prefix}_question1_options')
                        o2 = data.get(f'{prefix}_question2_options')
                        if q1 is None and q2 is None and o1 is None and o2 is None:
                            # Use existing JSON
                            raw = existing.get({
                                'logical': 'logical_reasoning_questions',
                                'numerical': 'numerical_ability_questions',
                                'verbal': 'verbal_ability_questions',
                                'spatial': 'spatial_reasoning_questions'
                            }[prefix])
                            try:
                                if isinstance(raw, str):
                                    return raw  # already JSON string
                                return _json.dumps(raw or [])
                            except Exception:
                                return _json.dumps([])
                        # Build from incoming
                        sec = []
                        if q1:
                            sec.append({'question': q1, 'options': o1 or []})
                        if q2:
                            sec.append({'question': q2, 'options': o2 or []})
                        return _json.dumps(sec)
                    cursor.execute(
                        """
                        UPDATE aptitude_tasks SET 
                            task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s,
                            logical_reasoning_questions=%s, numerical_ability_questions=%s, verbal_ability_questions=%s, spatial_reasoning_questions=%s,
                            instructions=%s, estimated_time=%s
                        WHERE id=%s
                        """,
                        (
                            existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'),
                            section_from('logical'), section_from('numerical'), section_from('verbal'), section_from('spatial'),
                            data.get('instructions', existing.get('instructions')), data.get('estimated_time', existing.get('estimated_time')),
                            task_id
                        )
                    )
                else:
                    cursor.execute(
                    """
                    UPDATE aptitude_tasks SET 
                        task_name=%s, age_min=%s, age_max=%s, difficulty_level=%s, instructions=%s, estimated_time=%s, example=%s,
                        logical_question1=%s, logical_question1_options=%s, logical_question2=%s, logical_question2_options=%s,
                        numerical_question1=%s, numerical_question1_options=%s, numerical_question2=%s, numerical_question2_options=%s,
                        verbal_question1=%s, verbal_question1_options=%s, verbal_question2=%s, verbal_question2_options=%s,
                        spatial_question1=%s, spatial_question1_options=%s, spatial_question2=%s, spatial_question2_options=%s
                    WHERE id=%s
                    """,
                    (
                        existing.get('task_name'), existing.get('age_min'), existing.get('age_max'), existing.get('difficulty_level'),
                        data.get('instructions', existing.get('instructions')), data.get('estimated_time', existing.get('estimated_time')), 
                        data.get('example', existing.get('example')),
                        data.get('logical_question1', existing.get('logical_question1')), _json.dumps(data.get('logical_question1_options', existing.get('logical_question1_options') or [])),
                        data.get('logical_question2', existing.get('logical_question2')), _json.dumps(data.get('logical_question2_options', existing.get('logical_question2_options') or [])),
                        data.get('numerical_question1', existing.get('numerical_question1')), _json.dumps(data.get('numerical_question1_options', existing.get('numerical_question1_options') or [])),
                        data.get('numerical_question2', existing.get('numerical_question2')), _json.dumps(data.get('numerical_question2_options', existing.get('numerical_question2_options') or [])),
                        data.get('verbal_question1', existing.get('verbal_question1')), _json.dumps(data.get('verbal_question1_options', existing.get('verbal_question1_options') or [])),
                        data.get('verbal_question2', existing.get('verbal_question2')), _json.dumps(data.get('verbal_question2_options', existing.get('verbal_question2_options') or [])),
                        data.get('spatial_question1', existing.get('spatial_question1')), _json.dumps(data.get('spatial_question1_options', existing.get('spatial_question1_options') or [])),
                        data.get('spatial_question2', existing.get('spatial_question2')), _json.dumps(data.get('spatial_question2_options', existing.get('spatial_question2_options') or [])),
                        task_id
                    )
                    )
            conn.commit()
            return jsonify({'success': True})
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        print(f"Admin category PUT/DELETE error: {e}")
        return jsonify({'success': False, 'message': 'Failed to update task'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()


@app.route('/api/admin/categories/<string:category_slug>/tasks/<int:task_id>', methods=['GET'])
def admin_category_task_get(category_slug: str, task_id: int):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    cfg = _category_config(category_slug)
    if not cfg:
        return jsonify({'success': False, 'message': 'Unknown category'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"SELECT * FROM {cfg['table']} WHERE id=%s", (task_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({'success': False, 'message': 'Task not found'}), 404
        return jsonify({'success': True, 'task': row})
    except Exception as e:
        print(f"Admin get single task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch task'}), 500


# Random task selection for Reading Aloud
@app.route('/api/reading-tasks-random/<int:user_id>', methods=['GET'])
def get_random_reading_task(user_id: int):
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        # Resolve age
        cursor.execute("SELECT age, date_of_birth FROM demographics WHERE user_id = %s", (user_id,))
        demo = cursor.fetchone()
        from datetime import datetime
        user_age = None
        if demo:
            if demo.get('age'):
                user_age = demo['age']
            elif demo.get('date_of_birth'):
                today = datetime.now()
                dob = demo['date_of_birth']
                user_age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        if not user_age:
            # fallback to any
            cursor.execute("SELECT id FROM reading_tasks ORDER BY RAND() LIMIT 1")
        else:
            cursor.execute("""
                SELECT id FROM reading_tasks WHERE age_min <= %s AND age_max >= %s ORDER BY RAND() LIMIT 1
            """, (user_age, user_age))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({'success': False, 'message': 'No reading tasks available'}), 404
        return jsonify({'success': True, 'task_id': row['id']})
    except Exception as e:
        print(f"Random reading task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get random task'}), 500

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        # Simple hardcoded admin credentials (replace with DB check in production)
        if username == 'admin' and password == 'admin123':
            session['is_admin'] = True
            # flash('Admin login successful!', 'success')
            return redirect(url_for('admin_portal'))
        else:
            flash('Invalid admin credentials', 'error')
            return render_template('admin_login.html')
    return render_template('admin_login.html')

# --- Task status API ---
@app.route('/api/user-tasks', methods=['GET'])
def get_user_tasks():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        query = "SELECT task_name, status FROM user_tasks WHERE user_id = %s"
        cursor.execute(query, (session['user_id'],))
        tasks = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'tasks': tasks})
    except Exception as e:
        print(f"Get user tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch tasks'}), 500

@app.route('/api/user-tasks', methods=['POST'])
def update_user_task():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    task_name = data.get('task_name')
    status = data.get('status')
    if not task_name or status not in ['Not Started', 'In Progress', 'Completed']:
        return jsonify({'success': False, 'message': 'Invalid task or status'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        # Upsert logic: update if exists, else insert
        query = """
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        """
        cursor.execute(query, (session['user_id'], task_name, status))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Task status updated'})
    except Exception as e:
        print(f"Update user task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to update task'}), 500

@app.route('/profile', methods=['GET'])
def profile():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    # Fetch user details from users and demographics
    user = None
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        query = '''
            SELECT u.name, u.email, d.date_of_birth, d.gender, d.native_language, d.education_level, d.dyslexia_status
            FROM users u
            LEFT JOIN demographics d ON u.id = d.user_id
            WHERE u.id = %s
        '''
        cursor.execute(query, (user_id,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"Profile fetch error: {e}")
    return render_template('profile.html', user=user)

@app.route('/api/profile', methods=['POST'])
def update_profile():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    dob = data.get('date_of_birth', '').strip()
    gender = data.get('gender', '').strip()
    native_language = data.get('native_language', '').strip()
    education_level = data.get('education_level', '').strip()
    dyslexia_status = data.get('dyslexia_status', '').strip()
    if not name or not email:
        return jsonify({'success': False, 'message': 'Name and email are required'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        # Update users table
        if password:
            import bcrypt
            password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            cursor.execute("UPDATE users SET name = %s, email = %s, password_hash = %s WHERE id = %s", (name, email, password_hash, user_id))
        else:
            cursor.execute("UPDATE users SET name = %s, email = %s WHERE id = %s", (name, email, user_id))
        # Update or insert demographics
        cursor.execute("SELECT id FROM demographics WHERE user_id = %s", (user_id,))
        if cursor.fetchone():
            cursor.execute("""
                UPDATE demographics SET date_of_birth = %s, gender = %s, native_language = %s, education_level = %s, dyslexia_status = %s WHERE user_id = %s
            """, (dob, gender, native_language, education_level, dyslexia_status, user_id))
        else:
            cursor.execute("""
                INSERT INTO demographics (user_id, date_of_birth, gender, native_language, education_level, dyslexia_status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (user_id, dob, gender, native_language, education_level, dyslexia_status))
        conn.commit()
        cursor.close()
        conn.close()
        session['email'] = email
        return jsonify({'success': True, 'message': 'Profile updated'})
    except Exception as e:
        print(f"Profile update error: {e}")
        return jsonify({'success': False, 'message': 'Failed to update profile'}), 500

@app.route('/task1.html')
def task1():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task1.html', user_id=user_id)

@app.route('/task2.html')
def task2():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task2.html', user_id=user_id)

@app.route('/task3.html')
def task3():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task3.html', user_id=user_id)

@app.route('/task4.html')
def task4():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task4.html', user_id=user_id)

@app.route('/task11.html')
def task11():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task11.html', user_id=user_id)

@app.route('/task22.html')
def task22():
    return render_template('task22.html')

@app.route('/task33.html')
def task33():
    return render_template('task33.html')

@app.route('/task44.html')
def task44():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task44.html', user_id=user_id)

@app.route('/task5.html')
def task5():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task5.html', user_id=user_id)

@app.route('/task55.html')
def task55():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task55.html', user_id=user_id)

@app.route('/task6.html')
def task6():
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('signin'))
    return render_template('task6.html', user_id=user_id)

@app.route('/aptitude.html')
def aptitude():
    # Require login to access the aptitude test so progress can be saved to a user
    if 'user_id' not in session:
        return redirect(url_for('signin'))
    return render_template('aptitude.html')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/api/save-progress', methods=['POST'])
def save_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    if 'audio' not in request.files:
        return jsonify({'success': False, 'message': 'No audio file provided'}), 400
    file = request.files['audio']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(f"user{session['user_id']}_" + datetime.now().strftime('%Y%m%d%H%M%S') + '_' + file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        # Save metadata to DB but don't mark task as completed
        try:
            conn = connect_db()
            cursor = conn.cursor()
            # Insert audio recording with progress status
            cursor.execute("""
                INSERT INTO audio_recordings (user_id, filename, task_name, uploaded_at)
                VALUES (%s, %s, %s, NOW())
            """, (session['user_id'], filename, 'Reading Aloud Task 1'))
            # Mark task as In Progress
            cursor.execute("""
                INSERT INTO user_tasks (user_id, task_name, status)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
            """, (session['user_id'], 'Reading Aloud Task 1', 'In Progress'))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Save progress DB error: {e}")
        return jsonify({'success': True, 'message': 'Progress saved successfully', 'filename': filename})
    else:
        return jsonify({'success': False, 'message': 'Invalid file type'}), 400

@app.route('/api/upload-audio', methods=['POST'])
def upload_audio():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    if 'audio' not in request.files:
        return jsonify({'success': False, 'message': 'No audio file provided'}), 400
    file = request.files['audio']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(f"user{session['user_id']}_" + datetime.now().strftime('%Y%m%d%H%M%S') + '_' + file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        # Save metadata to DB and mark task as completed
        try:
            conn = connect_db()
            cursor = conn.cursor()
            # Insert audio recording
            cursor.execute("""
                INSERT INTO audio_recordings (user_id, filename, uploaded_at)
                VALUES (%s, %s, NOW())
            """, (session['user_id'], filename))
            # Mark task as completed
            cursor.execute("""
                INSERT INTO user_tasks (user_id, task_name, status)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
            """, (session['user_id'], 'Reading Aloud Task 1', 'Completed'))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Audio metadata DB error: {e}")
        return jsonify({'success': True, 'message': 'Audio uploaded successfully and task marked as completed', 'filename': filename})
    else:
        return jsonify({'success': False, 'message': 'Invalid file type'}), 400

@app.route('/api/get-saved-progress', methods=['GET'])
def get_saved_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        # Get the most recent saved audio for this user and task
        cursor.execute("""
            SELECT filename FROM audio_recordings 
            WHERE user_id = %s AND task_name = %s 
            ORDER BY uploaded_at DESC 
            LIMIT 1
        """, (session['user_id'], 'Reading Aloud Task 1'))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if result:
            audio_url = f"/uploads/{result['filename']}"
            return jsonify({'success': True, 'saved_audio': audio_url})
        else:
            return jsonify({'success': True, 'saved_audio': None})
    except Exception as e:
        print(f"Get saved progress error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get saved progress'}), 500

@app.route('/api/save-typing-progress', methods=['POST'])
def save_typing_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    text = data.get('text', '')
    keystrokes = data.get('keystrokes', '')
    timer = data.get('timer', '')
    task_name = data.get('task_name', 'Typing Task')
    if not text:
        return jsonify({'success': False, 'message': 'No text provided'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        # Insert or update typing progress
        cursor.execute('''
            INSERT INTO typing_progress (user_id, text, keystrokes, timer, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE text=VALUES(text), keystrokes=VALUES(keystrokes), timer=VALUES(timer), updated_at=NOW()
        ''', (session['user_id'], text, keystrokes, timer))
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], 'Typing Task', 'In Progress'))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Progress saved successfully'})
    except Exception as e:
        print(f"Save typing progress DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to save progress'}), 500

@app.route('/api/submit-typing-task', methods=['POST'])
def submit_typing_task():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    text = data.get('text', '')
    keystrokes = data.get('keystrokes', '')
    timer = data.get('timer', '')
    task_name = data.get('task_name', 'Typing Task')
    if not text:
        return jsonify({'success': False, 'message': 'No text provided'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor()
        # Insert or update typing progress as completed
        cursor.execute('''
            INSERT INTO typing_progress (user_id, text, keystrokes, timer, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE text=VALUES(text), keystrokes=VALUES(keystrokes), timer=VALUES(timer), updated_at=NOW()
        ''', (session['user_id'], text, keystrokes, timer))
        # Mark task as Completed
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], task_name, 'Completed'))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Task submitted and marked as completed'})
    except Exception as e:
        print(f"Submit typing task DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to submit task'}), 500

@app.route('/api/get-typing-progress', methods=['GET'])
def get_typing_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('''
            SELECT text, keystrokes, timer FROM typing_progress WHERE user_id = %s ORDER BY updated_at DESC LIMIT 1
        ''', (session['user_id'],))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        if result:
            return jsonify({'success': True, 'progress': result})
        else:
            return jsonify({'success': True, 'progress': None})
    except Exception as e:
        print(f"Get typing progress error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get progress'}), 500

@app.route('/api/save-comprehension-progress', methods=['POST'])
def save_comprehension_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    q1 = data.get('q1', '')
    q2 = data.get('q2', '')
    q3 = data.get('q3', '')
    task_name = data.get('task_name', 'Reading Comprehension')
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO comprehension_progress (user_id, q1, q2, q3, status, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE q1=VALUES(q1), q2=VALUES(q2), q3=VALUES(q3), status=VALUES(status), updated_at=NOW()
        ''', (session['user_id'], q1, q2, q3, 'In Progress'))
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], task_name, 'In Progress'))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Progress saved successfully'})
    except Exception as e:
        print(f"Save comprehension progress DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to save progress'}), 500

@app.route('/api/get-comprehension-progress', methods=['GET'])
def get_comprehension_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('''
            SELECT q1, q2, q3, status FROM comprehension_progress WHERE user_id = %s
        ''', (session['user_id'],))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        if result:
            return jsonify({'success': True, 'progress': result})
        else:
            return jsonify({'success': True, 'progress': None})
    except Exception as e:
        print(f"Get comprehension progress error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get progress'}), 500

@app.route('/api/submit-comprehension', methods=['POST'])
def submit_comprehension():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    q1 = data.get('q1', '')
    q2 = data.get('q2', '')
    q3 = data.get('q3', '')
    task_name = data.get('task_name', 'Reading Comprehension')
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO comprehension_progress (user_id, q1, q2, q3, status, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE q1=VALUES(q1), q2=VALUES(q2), q3=VALUES(q3), status=VALUES(status), updated_at=NOW()
        ''', (session['user_id'], q1, q2, q3, 'Completed'))
        # Mark task as Completed
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], task_name, 'Completed'))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Task submitted and marked as completed'})
    except Exception as e:
        print(f"Submit comprehension DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to submit task'}), 500


@app.route('/api/save-aptitude-progress', methods=['POST'])
def save_aptitude_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    logical_reasoning_score = data.get('logical_reasoning_score', 0)
    numerical_ability_score = data.get('numerical_ability_score', 0)
    verbal_ability_score = data.get('verbal_ability_score', 0)
    spatial_reasoning_score = data.get('spatial_reasoning_score', 0)
    total_score = data.get('total_score', 0)
    # New detailed progress fields
    answers = data.get('answers')  # expected dict of {section: {questionIndex: value}}
    current_section = data.get('current_section')
    answered_count = data.get('answered_count', 0)
    progress_percent = data.get('progress_percent', 0)
    
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # Ensure schema has columns to store detailed progress (idempotent on MySQL 8+)
        try:
            cursor.execute(
                """
                ALTER TABLE aptitude_progress
                ADD COLUMN IF NOT EXISTS answers JSON NULL,
                ADD COLUMN IF NOT EXISTS current_section VARCHAR(50) NULL,
                ADD COLUMN IF NOT EXISTS answered_count INT DEFAULT 0,
                ADD COLUMN IF NOT EXISTS progress_percent INT DEFAULT 0
                """
            )
            conn.commit()
        except Exception as _e:
            # Ignore if DB doesn't support IF NOT EXISTS or columns already exist
            pass

        # Insert/Update progress with answers/state
        cursor.execute('''
            INSERT INTO aptitude_progress (
                user_id,
                logical_reasoning_score,
                numerical_ability_score,
                verbal_ability_score,
                spatial_reasoning_score,
                total_score,
                status,
                answers,
                current_section,
                answered_count,
                progress_percent,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE 
                logical_reasoning_score=VALUES(logical_reasoning_score), 
                numerical_ability_score=VALUES(numerical_ability_score), 
                verbal_ability_score=VALUES(verbal_ability_score), 
                spatial_reasoning_score=VALUES(spatial_reasoning_score), 
                total_score=VALUES(total_score), 
                status=VALUES(status), 
                answers=VALUES(answers),
                current_section=VALUES(current_section),
                answered_count=VALUES(answered_count),
                progress_percent=VALUES(progress_percent),
                updated_at=NOW()
        ''', (
            session['user_id'],
            logical_reasoning_score,
            numerical_ability_score,
            verbal_ability_score,
            spatial_reasoning_score,
            total_score,
            'In Progress',
            json.dumps(answers) if answers is not None else None,
            current_section,
            answered_count,
            progress_percent
        ))
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], 'Aptitude Test', 'In Progress'))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Progress saved successfully'})
    except Exception as e:
        print(f"Save aptitude progress DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to save progress'}), 500


@app.route('/api/get-aptitude-progress', methods=['GET'])
def get_aptitude_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        # Try to ensure extended columns exist; ignore errors on older MySQL
        try:
            cursor.execute(
                """
                ALTER TABLE aptitude_progress
                ADD COLUMN IF NOT EXISTS answers JSON NULL,
                ADD COLUMN IF NOT EXISTS current_section VARCHAR(50) NULL,
                ADD COLUMN IF NOT EXISTS answered_count INT DEFAULT 0,
                ADD COLUMN IF NOT EXISTS progress_percent INT DEFAULT 0
                """
            )
            conn.commit()
        except Exception:
            pass

        # Attempt full select including extended columns; fall back if unavailable
        try:
            cursor.execute('''
                SELECT 
                    logical_reasoning_score,
                    numerical_ability_score,
                    verbal_ability_score,
                    spatial_reasoning_score,
                    total_score,
                    status,
                    answers,
                    current_section,
                    answered_count,
                    progress_percent
                FROM aptitude_progress WHERE user_id = %s
            ''', (session['user_id'],))
            result = cursor.fetchone()
        except Exception:
            cursor.close()
            cursor = conn.cursor(dictionary=True)
            cursor.execute('''
                SELECT 
                    logical_reasoning_score,
                    numerical_ability_score,
                    verbal_ability_score,
                    spatial_reasoning_score,
                    total_score,
                    status
                FROM aptitude_progress WHERE user_id = %s
            ''', (session['user_id'],))
            base = cursor.fetchone()
            if base:
                # Provide default None/0s for missing fields
                base['answers'] = None
                base['current_section'] = None
                base['answered_count'] = 0
                base['progress_percent'] = 0
            result = base
        cursor.close()
        conn.close()
        if result:
            return jsonify({'success': True, 'progress': result})
        else:
            return jsonify({'success': True, 'progress': None})
    except Exception as e:
        print(f"Get aptitude progress error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get progress'}), 500


@app.route('/api/submit-aptitude', methods=['POST'])
def submit_aptitude():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    data = request.get_json()
    logical_reasoning_score = data.get('logical_reasoning_score', 0)
    numerical_ability_score = data.get('numerical_ability_score', 0)
    verbal_ability_score = data.get('verbal_ability_score', 0)
    spatial_reasoning_score = data.get('spatial_reasoning_score', 0)
    total_score = data.get('total_score', 0)
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO aptitude_progress (user_id, logical_reasoning_score, numerical_ability_score, verbal_ability_score, spatial_reasoning_score, total_score, status, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE 
                logical_reasoning_score=VALUES(logical_reasoning_score), 
                numerical_ability_score=VALUES(numerical_ability_score), 
                verbal_ability_score=VALUES(verbal_ability_score), 
                spatial_reasoning_score=VALUES(spatial_reasoning_score), 
                total_score=VALUES(total_score), 
                status=VALUES(status), 
                updated_at=NOW()
        ''', (session['user_id'], logical_reasoning_score, numerical_ability_score, verbal_ability_score, spatial_reasoning_score, total_score, 'Completed'))
        # Mark task as Completed
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], 'Aptitude Test', 'Completed'))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Task submitted and marked as completed'})
    except Exception as e:
        print(f"Submit aptitude DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to submit task'}), 500


@app.route('/api/reading-tasks/<int:user_id>', methods=['GET'])
def get_reading_tasks(user_id):
    """Get age-appropriate reading tasks for a user"""
    try:
        print(f"Fetching reading tasks for user_id: {user_id}")
        
        conn = connect_db()
        if not conn:
            print("Database connection failed")
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # First check if user exists
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        user_exists = cursor.fetchone()
        if not user_exists:
            print(f"User {user_id} not found")
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's age from demographics
        cursor.execute("SELECT age, date_of_birth FROM demographics WHERE user_id = %s", (user_id,))
        result = cursor.fetchone()
        
        print(f"Demographics result: {result}")
        
        if not result:
            print(f"No demographics found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultReadingTasks()
        
        if not result['age'] and not result['date_of_birth']:
            print(f"No age or date_of_birth found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultReadingTasks()
        
        # Use age if available, otherwise calculate from date_of_birth
        if result['age']:
            user_age = result['age']
        elif result['date_of_birth']:
            from datetime import datetime
            today = datetime.now()
            birth_date = result['date_of_birth']
            user_age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        else:
            user_age = None
        
        print(f"Calculated user age: {user_age}")
        
        if user_age is None or user_age <= 0:
            return jsonify({'success': False, 'message': 'Invalid age. Please update your profile.'}), 400
        
        # Check if reading_tasks table exists
        cursor.execute("SHOW TABLES LIKE 'reading_tasks'")
        table_exists = cursor.fetchone()
        if not table_exists:
            print("reading_tasks table does not exist")
            return jsonify({'success': False, 'message': 'Reading tasks not configured. Please contact administrator.'}), 500
        
        # Get appropriate reading tasks for user's age
        cursor.execute("""
            SELECT * FROM reading_tasks 
            WHERE age_min <= %s AND age_max >= %s 
            ORDER BY difficulty_level, age_min
        """, (user_age, user_age))
        
        tasks = cursor.fetchall()
        print(f"Found {len(tasks)} tasks for age {user_age}")
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': f'No reading tasks found for age {user_age}. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': user_age,
            'total_tasks': len(tasks)
        })
        
    except Exception as e:
        print(f"Get reading tasks error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Failed to fetch reading tasks: {str(e)}'}), 500


def getDefaultReadingTasks():
    """Return default reading tasks when user age is not available"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get all available reading tasks
        cursor.execute("SELECT * FROM reading_tasks ORDER BY difficulty_level, age_min")
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': 'No reading tasks available. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': 'Not specified',
            'total_tasks': len(tasks),
            'message': 'Showing all available reading tasks. Please complete your profile setup for personalized tasks.'
        })
        
    except Exception as e:
        print(f"Get default reading tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch default reading tasks'}), 500


@app.route('/api/reading-task/<int:task_id>', methods=['GET'])
def get_reading_task_by_id(task_id):
    """Get a specific reading task by ID"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get the specific reading task
        cursor.execute("SELECT * FROM reading_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not task:
            return jsonify({'success': False, 'message': 'Reading task not found'}), 404
        
        return jsonify({
            'success': True, 
            'task': task
        })
        
    except Exception as e:
        print(f"Get reading task by ID error: {e}")
        return jsonify({'success': False, 'message': f'Failed to fetch reading task: {str(e)}'}), 500


@app.route('/api/reading-comprehension-tasks/<int:user_id>', methods=['GET'])
def get_reading_comprehension_tasks(user_id):
    """Get age-appropriate reading comprehension tasks for a user"""
    try:
        print(f"Fetching reading comprehension tasks for user_id: {user_id}")
        
        conn = connect_db()
        if not conn:
            print("Database connection failed")
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # First check if user exists
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        user_exists = cursor.fetchone()
        if not user_exists:
            print(f"User {user_id} not found")
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's age from demographics
        cursor.execute("SELECT age, date_of_birth FROM demographics WHERE user_id = %s", (user_id,))
        result = cursor.fetchone()
        
        print(f"Demographics result: {result}")
        
        if not result:
            print(f"No demographics found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultReadingComprehensionTasks()
        
        if not result['age'] and not result['date_of_birth']:
            print(f"No age or date_of_birth found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultReadingComprehensionTasks()
        
        # Use age if available, otherwise calculate from date_of_birth
        if result['age']:
            user_age = result['age']
        elif result['date_of_birth']:
            from datetime import datetime
            today = datetime.now()
            birth_date = result['date_of_birth']
            user_age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        else:
            user_age = None
        
        print(f"Calculated user age: {user_age}")
        
        if user_age is None or user_age <= 0:
            return jsonify({'success': False, 'message': 'Invalid age. Please update your profile.'}), 400
        
        # Check if reading_comprehension_tasks table exists
        cursor.execute("SHOW TABLES LIKE 'reading_comprehension_tasks'")
        table_exists = cursor.fetchone()
        if not table_exists:
            print("reading_comprehension_tasks table does not exist")
            return jsonify({'success': False, 'message': 'Reading comprehension tasks not configured. Please contact administrator.'}), 500
        
        # Get appropriate reading comprehension tasks for user's age
        cursor.execute("""
            SELECT * FROM reading_comprehension_tasks 
            WHERE age_min <= %s AND age_max >= %s 
            ORDER BY difficulty_level, age_min
        """, (user_age, user_age))
        
        tasks = cursor.fetchall()
        print(f"Found {len(tasks)} comprehension tasks for age {user_age}")
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': f'No reading comprehension tasks found for age {user_age}. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': user_age,
            'total_tasks': len(tasks)
        })
        
    except Exception as e:
        print(f"Get reading comprehension tasks error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Failed to fetch reading comprehension tasks: {str(e)}'}), 500


def getDefaultReadingComprehensionTasks():
    """Return default reading comprehension tasks when user age is not available"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get all available reading comprehension tasks
        cursor.execute("SELECT * FROM reading_comprehension_tasks ORDER BY difficulty_level, age_min")
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': 'No reading comprehension tasks available. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': 'Not specified',
            'total_tasks': len(tasks),
            'message': 'Showing all available reading comprehension tasks. Please complete your profile setup for personalized tasks.'
        })
        
    except Exception as e:
        print(f"Get default reading comprehension tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch default reading comprehension tasks'}), 500


@app.route('/api/reading-comprehension-task/<int:task_id>', methods=['GET'])
def get_reading_comprehension_task_by_id(task_id):
    """Get a specific reading comprehension task by ID"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get the specific reading comprehension task
        cursor.execute("SELECT * FROM reading_comprehension_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not task:
            return jsonify({'success': False, 'message': 'Reading comprehension task not found'}), 404
        
        return jsonify({
            'success': True, 
            'task': task
        })
        
    except Exception as e:
        print(f"Get reading comprehension task by ID error: {e}")
        return jsonify({'success': False, 'message': f'Failed to fetch reading comprehension task: {str(e)}'}), 500


@app.route('/api/typing-tasks/<int:user_id>', methods=['GET'])
def get_typing_tasks(user_id):
    """Get age-appropriate typing tasks for a user"""
    try:
        print(f"Fetching typing tasks for user_id: {user_id}")
        
        conn = connect_db()
        if not conn:
            print("Database connection failed")
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # First check if user exists
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        user_exists = cursor.fetchone()
        if not user_exists:
            print(f"User {user_id} not found")
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's age from demographics
        cursor.execute("SELECT age, date_of_birth FROM demographics WHERE user_id = %s", (user_id,))
        result = cursor.fetchone()
        
        print(f"Demographics result: {result}")
        
        if not result:
            print(f"No demographics found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultTypingTasks()
        
        if not result['age'] and not result['date_of_birth']:
            print(f"No age or date_of_birth found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultTypingTasks()
        
        # Use age if available, otherwise calculate from date_of_birth
        if result['age']:
            user_age = result['age']
        elif result['date_of_birth']:
            from datetime import datetime
            today = datetime.now()
            birth_date = result['date_of_birth']
            user_age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        else:
            user_age = None
        
        print(f"Calculated user age: {user_age}")
        
        if user_age is None or user_age <= 0:
            return jsonify({'success': False, 'message': 'Invalid age. Please update your profile.'}), 400
        
        # Check if typing_tasks table exists
        cursor.execute("SHOW TABLES LIKE 'typing_tasks'")
        table_exists = cursor.fetchone()
        if not table_exists:
            print("typing_tasks table does not exist")
            return jsonify({'success': False, 'message': 'Typing tasks not configured. Please contact administrator.'}), 500
        
        # Get appropriate typing tasks for user's age
        cursor.execute("""
            SELECT * FROM typing_tasks 
            WHERE age_min <= %s AND age_max >= %s 
            ORDER BY difficulty_level, age_min
        """, (user_age, user_age))
        
        tasks = cursor.fetchall()
        print(f"Found {len(tasks)} typing tasks for age {user_age}")
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': f'No typing tasks found for age {user_age}. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': user_age,
            'total_tasks': len(tasks)
        })
        
    except Exception as e:
        print(f"Get typing tasks error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Failed to fetch typing tasks: {str(e)}'}), 500


def getDefaultTypingTasks():
    """Return default typing tasks when user age is not available"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get all available typing tasks
        cursor.execute("SELECT * FROM typing_tasks ORDER BY difficulty_level, age_min")
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': 'No typing tasks available. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': 'Not specified',
            'total_tasks': len(tasks),
            'message': 'Showing all available typing tasks. Please complete your profile setup for personalized tasks.'
        })
        
    except Exception as e:
        print(f"Get default typing tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch default typing tasks'}), 500


@app.route('/api/typing-task/<int:task_id>', methods=['GET'])
def get_typing_task_by_id(task_id):
    """Get a specific typing task by ID"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get the specific typing task
        cursor.execute("SELECT * FROM typing_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not task:
            return jsonify({'success': False, 'message': 'Typing task not found'}), 404
        
        return jsonify({
            'success': True, 
            'task': task
        })
        
    except Exception as e:
        print(f"Get typing task by ID error: {e}")
        return jsonify({'success': False, 'message': f'Failed to fetch typing task: {str(e)}'}), 500


@app.route('/api/start-typing-task', methods=['POST'])
def start_typing_task():
    """Mark a typing task as started"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    task_name = data.get('task_name')
    
    if not task_id or not task_name:
        return jsonify({'success': False, 'message': 'Task information required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], f'Typing Task {task_name}', 'In Progress'))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Task started successfully'})
        
    except Exception as e:
        print(f"Start typing task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to start task'}), 500


@app.route('/api/start-reading-comprehension-task', methods=['POST'])
def start_reading_comprehension_task():
    """Mark a reading comprehension task as started"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    task_name = data.get('task_name')
    
    if not task_id or not task_name:
        return jsonify({'success': False, 'message': 'Task information required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], f'Reading Comprehension Task {task_name}', 'In Progress'))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Task started successfully'})
        
    except Exception as e:
        print(f"Start reading comprehension task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to start task'}), 500


@app.route('/api/start-reading-task', methods=['POST'])
def start_reading_task():
    """Mark a reading task as started"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    task_name = data.get('task_name')
    
    if not task_id or not task_name:
        return jsonify({'success': False, 'message': 'Task information required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], f'Reading Task {task_name}', 'In Progress'))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Task started successfully'})
        
    except Exception as e:
        print(f"Start reading task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to start task'}), 500


@app.route('/api/get-task-status', methods=['GET'])
def get_task_status():
    """Get the current status of a specific task for the logged-in user"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    task_name = request.args.get('task_name')
    if not task_name:
        return jsonify({'success': False, 'message': 'Task name required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        
        # Get the task status for the current user
        cursor.execute('''
            SELECT status FROM user_tasks 
            WHERE user_id = %s AND task_name = %s
        ''', (session['user_id'], task_name))
        
        result = cursor.fetchone()
        status = result['status'] if result else None

        # Backward-compatibility fallback for older records
        # Some older endpoints saved mathematical comprehension as a generic name
        # without the specific task suffix. If no exact match is found, check
        # the generic record to preserve users' in-progress state.
        if status is None and task_name.startswith('Mathematical Comprehension Task '):
            cursor.execute('''
                SELECT status FROM user_tasks 
                WHERE user_id = %s AND task_name = %s
            ''', (session['user_id'], 'Mathematical Comprehension'))
            fallback = cursor.fetchone()
            status = fallback['status'] if fallback else 'Not Started'
        elif status is None:
            status = 'Not Started'
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'status': status})
        
    except Exception as e:
        print(f"Get task status error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get task status'}), 500


@app.route('/api/mathematical-comprehension-tasks/<int:user_id>', methods=['GET'])
def get_mathematical_comprehension_tasks(user_id):
    """Get age-appropriate mathematical comprehension tasks for a user"""
    try:
        print(f"Fetching mathematical comprehension tasks for user_id: {user_id}")
        
        conn = connect_db()
        if not conn:
            print("Database connection failed")
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # First check if user exists
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        user_exists = cursor.fetchone()
        if not user_exists:
            print(f"User {user_id} not found")
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's age from demographics
        cursor.execute("SELECT age, date_of_birth FROM demographics WHERE user_id = %s", (user_id,))
        result = cursor.fetchone()
        
        print(f"Demographics result: {result}")
        
        if not result:
            print(f"No demographics found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultMathematicalComprehensionTasks()
        
        if not result['age'] and not result['date_of_birth']:
            print(f"No age or date_of_birth found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultMathematicalComprehensionTasks()
        
        # Use age if available, otherwise calculate from date_of_birth
        if result['age']:
            user_age = result['age']
        elif result['date_of_birth']:
            from datetime import datetime
            today = datetime.now()
            birth_date = result['date_of_birth']
            user_age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        else:
            user_age = None
        
        print(f"Calculated user age: {user_age}")
        
        if user_age is None or user_age <= 0:
            return jsonify({'success': False, 'message': 'Invalid age. Please update your profile.'}), 400
        
        # Check if mathematical_comprehension_tasks table exists
        cursor.execute("SHOW TABLES LIKE 'mathematical_comprehension_tasks'")
        table_exists = cursor.fetchone()
        if not table_exists:
            print("mathematical_comprehension_tasks table does not exist")
            return jsonify({'success': False, 'message': 'Mathematical comprehension tasks not configured. Please contact administrator.'}), 500
        
        # Get appropriate mathematical comprehension tasks for user's age
        cursor.execute("""
            SELECT * FROM mathematical_comprehension_tasks 
            WHERE age_min <= %s AND age_max >= %s 
            ORDER BY difficulty_level, age_min
        """, (user_age, user_age))
        
        tasks = cursor.fetchall()
        print(f"Found {len(tasks)} mathematical comprehension tasks for age {user_age}")
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': f'No mathematical comprehension tasks found for age {user_age}. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': user_age,
            'total_tasks': len(tasks)
        })
        
    except Exception as e:
        print(f"Get mathematical comprehension tasks error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Failed to fetch mathematical comprehension tasks: {str(e)}'}), 500


def getDefaultMathematicalComprehensionTasks():
    """Return default mathematical comprehension tasks when user age is not available"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get all available mathematical comprehension tasks
        cursor.execute("SELECT * FROM mathematical_comprehension_tasks ORDER BY difficulty_level, age_min")
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': 'No mathematical comprehension tasks available. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': 'Not specified',
            'total_tasks': len(tasks),
            'message': 'Showing all available mathematical comprehension tasks. Please complete your profile setup for personalized tasks.'
        })
        
    except Exception as e:
        print(f"Get default mathematical comprehension tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch default mathematical comprehension tasks'}), 500


@app.route('/api/mathematical-comprehension-task/<int:task_id>', methods=['GET'])
def get_mathematical_comprehension_task_by_id(task_id):
    """Get a specific mathematical comprehension task by ID"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get the specific mathematical comprehension task
        cursor.execute("SELECT * FROM mathematical_comprehension_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not task:
            return jsonify({'success': False, 'message': 'Mathematical comprehension task not found'}), 404
        
        return jsonify({
            'success': True, 
            'task': task
        })
        
    except Exception as e:
        print(f"Get mathematical comprehension task by ID error: {e}")
        return jsonify({'success': False, 'message': f'Failed to fetch mathematical comprehension task: {str(e)}'}), 500


@app.route('/api/start-mathematical-comprehension-task', methods=['POST'])
def start_mathematical_comprehension_task():
    """Mark a mathematical comprehension task as started"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    task_name = data.get('task_name')
    
    if not task_id or not task_name:
        return jsonify({'success': False, 'message': 'Task information required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], f'Mathematical Comprehension Task {task_name}', 'In Progress'))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Task started successfully'})
        
    except Exception as e:
        print(f"Start mathematical comprehension task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to start task'}), 500

@app.route('/api/writing-tasks/<int:user_id>', methods=['GET'])
def get_writing_tasks(user_id):
    """Get age-appropriate writing tasks for a user"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get user's age from demographics table
        cursor.execute("SELECT age FROM demographics WHERE user_id = %s", (user_id,))
        age_result = cursor.fetchone()
        
        user_age = None
        if age_result and age_result['age']:
            user_age = age_result['age']
        
        # Get writing tasks based on age
        if user_age:
            cursor.execute("""
                SELECT * FROM writing_tasks 
                WHERE age_min <= %s AND age_max >= %s 
                ORDER BY difficulty_level, age_min
            """, (user_age, user_age))
            tasks = cursor.fetchall()
            
            if tasks:
                return jsonify({
                    'success': True,
                    'tasks': tasks,
                    'user_age': user_age,
                    'message': f'Showing writing tasks for age {user_age}'
                })
            else:
                # If no tasks found for specific age, get default tasks
                return getDefaultWritingTasks()
        else:
            # If no age found, get default tasks
            return getDefaultWritingTasks()
            
    except Exception as e:
        print(f"Get writing tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to load writing tasks'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

def getDefaultWritingTasks():
    """Get default writing tasks when user age is not available"""
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        
        # Get all writing tasks ordered by difficulty
        cursor.execute("""
            SELECT * FROM writing_tasks 
            ORDER BY difficulty_level, age_min
        """)
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'tasks': tasks,
            'user_age': 'Not specified',
            'message': 'Showing all available writing tasks'
        })
        
    except Exception as e:
        print(f"Get default writing tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to load writing tasks'}), 500

@app.route('/api/writing-task/<int:task_id>', methods=['GET'])
def get_writing_task_by_id(task_id):
    """Get a specific writing task by ID"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get the specific writing task
        cursor.execute("SELECT * FROM writing_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not task:
            return jsonify({'success': False, 'message': 'Writing task not found'}), 404
        
        return jsonify({
            'success': True, 
            'task': task
        })
        
    except Exception as e:
        print(f"Get writing task by ID error: {e}")
        return jsonify({'success': False, 'message': 'Failed to load writing task'}), 500

@app.route('/api/start-writing-task', methods=['POST'])
def start_writing_task():
    """Mark a writing task as started"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    task_name = data.get('task_name')
    
    if not task_id or not task_name:
        return jsonify({'success': False, 'message': 'Task information required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], f'Writing Task {task_name}', 'In Progress'))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Task started successfully'})
        
    except Exception as e:
        print(f"Start writing task DB error: {e}")
        return jsonify({'success': False, 'message': 'Failed to start task'}), 500

@app.route('/api/aptitude-tasks/<int:user_id>', methods=['GET'])
def get_aptitude_tasks(user_id):
    """Get age-appropriate aptitude tasks for a user"""
    try:
        print(f"Fetching aptitude tasks for user_id: {user_id}")
        
        conn = connect_db()
        if not conn:
            print("Database connection failed")
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # First check if user exists
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        user_exists = cursor.fetchone()
        if not user_exists:
            print(f"User {user_id} not found")
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get user's age from demographics
        cursor.execute("SELECT age, date_of_birth FROM demographics WHERE user_id = %s", (user_id,))
        result = cursor.fetchone()
        
        print(f"Demographics result: {result}")
        
        if not result:
            print(f"No demographics found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultAptitudeTasks()
        
        if not result['age'] and not result['date_of_birth']:
            print(f"No age or date_of_birth found for user {user_id}")
            # Return default tasks instead of error
            return getDefaultAptitudeTasks()
        
        # Use age if available, otherwise calculate from date_of_birth
        if result['age']:
            user_age = result['age']
        elif result['date_of_birth']:
            from datetime import datetime
            today = datetime.now()
            birth_date = result['date_of_birth']
            user_age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        else:
            user_age = None
        
        print(f"Calculated user age: {user_age}")
        
        if user_age is None or user_age <= 0:
            return jsonify({'success': False, 'message': 'Invalid age. Please update your profile.'}), 400
        
        # Check if aptitude_tasks table exists
        cursor.execute("SHOW TABLES LIKE 'aptitude_tasks'")
        table_exists = cursor.fetchone()
        if not table_exists:
            print("aptitude_tasks table does not exist")
            return jsonify({'success': False, 'message': 'Aptitude tasks not configured. Please contact administrator.'}), 500
        
        # Get appropriate aptitude tasks for user's age
        cursor.execute("""
            SELECT * FROM aptitude_tasks 
            WHERE age_min <= %s AND age_max >= %s 
            ORDER BY difficulty_level, age_min
        """, (user_age, user_age))
        
        tasks = cursor.fetchall()
        print(f"Found {len(tasks)} aptitude tasks for age {user_age}")
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': f'No aptitude tasks found for age {user_age}. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': user_age,
            'total_tasks': len(tasks)
        })
        
    except Exception as e:
        print(f"Get aptitude tasks error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Failed to fetch aptitude tasks: {str(e)}'}), 500


def getDefaultAptitudeTasks():
    """Return default aptitude tasks when user age is not available"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get all available aptitude tasks
        cursor.execute("SELECT * FROM aptitude_tasks ORDER BY difficulty_level, age_min")
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        if not tasks:
            return jsonify({'success': False, 'message': 'No aptitude tasks available. Please contact administrator.'}), 404
        
        return jsonify({
            'success': True, 
            'tasks': tasks, 
            'user_age': 'Not specified',
            'total_tasks': len(tasks),
            'message': 'Showing all available aptitude tasks. Please complete your profile setup for personalized tasks.'
        })
        
    except Exception as e:
        print(f"Get default aptitude tasks error: {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch default aptitude tasks'}), 500


@app.route('/api/aptitude-task/<int:task_id>', methods=['GET'])
def get_aptitude_task_by_id(task_id):
    """Get a specific aptitude task by ID"""
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get the specific aptitude task
        cursor.execute("SELECT * FROM aptitude_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not task:
            return jsonify({'success': False, 'message': 'Aptitude task not found'}), 404
        
        return jsonify({
            'success': True, 
            'task': task
        })
        
    except Exception as e:
        print(f"Get aptitude task by ID error: {e}")
        return jsonify({'success': False, 'message': f'Failed to fetch aptitude task: {str(e)}'}), 500


@app.route('/api/start-aptitude-task', methods=['POST'])
def start_aptitude_task():
    """Mark an aptitude task as started"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    task_name = data.get('task_name')
    
    if not task_id or not task_name:
        return jsonify({'success': False, 'message': 'Task information required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor()
        
        # Mark task as In Progress
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (session['user_id'], f'Aptitude Test {task_name}', 'In Progress'))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Task started successfully'})
        
    except Exception as e:
        print(f"Start aptitude task error: {e}")
        return jsonify({'success': False, 'message': 'Failed to start task'}), 500

@app.route('/api/upload-writing', methods=['POST'])
def upload_writing():
    """Upload writing sample image"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    if 'writing_image' not in request.files:
        return jsonify({'success': False, 'message': 'No image file provided'}), 400
    
    file = request.files['writing_image']
    task_id = request.form.get('task_id')
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400
    
    if not task_id:
        return jsonify({'success': False, 'message': 'Task ID required'}), 400
    
    # Check if file is an image
    if file and allowed_file(file.filename):
        filename = secure_filename(f"writing_user{session['user_id']}_task{task_id}_" + datetime.now().strftime('%Y%m%d%H%M%S') + '_' + file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Save writing sample to database
            cursor.execute("""
                INSERT INTO writing_samples (user_id, task_id, filename, status)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE filename = VALUES(filename), status = VALUES(status), uploaded_at = NOW()
            """, (session['user_id'], task_id, filename, 'Completed'))
            
            # Mark task as completed
            cursor.execute("""
                INSERT INTO user_tasks (user_id, task_name, status)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
            """, (session['user_id'], 'Writing Task', 'Completed'))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Writing sample uploaded successfully', 'filename': filename})
            
        except Exception as e:
            print(f"Upload writing DB error: {e}")
            return jsonify({'success': False, 'message': 'Failed to save writing sample'}), 500
    else:
        return jsonify({'success': False, 'message': 'Invalid file type. Please upload an image.'}), 400

@app.route('/api/save-writing-progress', methods=['POST'])
def save_writing_progress():
    """Save writing task progress without marking as completed"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    
    if 'writing_image' not in request.files:
        return jsonify({'success': False, 'message': 'No image file provided'}), 400
    
    file = request.files['writing_image']
    task_id = request.form.get('task_id')
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400
    
    if not task_id:
        return jsonify({'success': False, 'message': 'Task ID required'}), 400
    
    # Check if file is an image
    if file and allowed_file(file.filename):
        filename = secure_filename(f"writing_user{session['user_id']}_task{task_id}_" + datetime.now().strftime('%Y%m%d%H%M%S') + '_' + file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Save writing sample to database with In Progress status
            cursor.execute("""
                INSERT INTO writing_samples (user_id, task_id, filename, status)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE filename = VALUES(filename), uploaded_at = NOW()
            """, (session['user_id'], task_id, filename, 'In Progress'))
            
            # Mark task as In Progress
            cursor.execute("""
                INSERT INTO user_tasks (user_id, task_name, status)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
            """, (session['user_id'], 'Writing Task', 'In Progress'))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Writing progress saved successfully', 'filename': filename})
            
        except Exception as e:
            print(f"Save writing progress DB error: {e}")
            return jsonify({'success': False, 'message': 'Failed to save writing progress'}), 500
    else:
        return jsonify({'success': False, 'message': 'Invalid file type. Please upload an image.'}), 400


@app.route('/api/admin/tasks', methods=['GET'])
def get_all_tasks():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM tasks ORDER BY id')
        tasks = cursor.fetchall()
        cursor.close()
        conn.close()
        
        print(f"Found {len(tasks)} tasks in database")
        for task in tasks:
            print(f"Task: {task}")
        
        return jsonify({'success': True, 'tasks': tasks})
    except Exception as e:
        print(f"Error in get_all_tasks: {e}")
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500

# @app.route('/api/admin/tasks', methods=['POST'])
# def add_task():
#     if not session.get('is_admin'):
#         return jsonify({'success': False, 'message': 'Unauthorized'}), 401
#     data = request.get_json()
#     task_name = data.get('task_name')
#     description = data.get('description', '')
#     instructions = data.get('instructions', '')
#     estimated_time = data.get('estimated_time', '')
#     devices_required = data.get('devices_required', '')
#     example = data.get('example', '')
#     if not task_name or not description or not instructions or not estimated_time or not devices_required or not example:
#         return jsonify({'success': False, 'message': 'All fields are required'}), 400
#     conn = connect_db()
#     cursor = conn.cursor()
#     try:
#         cursor.execute('''
#             INSERT INTO tasks (task_name, description, instructions, estimated_time, devices_required, example)
#             VALUES (%s, %s, %s, %s, %s, %s)
#         ''', (task_name, description, instructions, estimated_time, devices_required, example))
#         conn.commit()
#         # Optionally: create the HTML file for the task here (see below)
#         return jsonify({'success': True, 'message': 'Task added'})
#     except Exception as e:
#         conn.rollback()
#         return jsonify({'success': False, 'message': str(e)}), 500
#     finally:
#         cursor.close()
#         conn.close()

@app.route('/api/admin/tasks', methods=['POST'])
def add_task():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.get_json()
    task_name = data.get('task_name')
    description = data.get('description', '')
    instructions = data.get('instructions', '')
    estimated_time = data.get('estimated_time', '')
    devices_required = data.get('devices_required', '')
    example = data.get('example', '')
    main_content = data.get('main_content', '')
    if not task_name or not description or not instructions or not estimated_time or not devices_required or not example:
        return jsonify({'success': False, 'message': 'All fields are required'}), 400
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO tasks (task_name, description, instructions, estimated_time, devices_required, example)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (task_name, description, instructions, estimated_time, devices_required, example))
        conn.commit()
        # Create the HTML file for the task
        create_task_html(task_name, instructions, estimated_time, devices_required, example, main_content)
        return jsonify({'success': True, 'message': 'Task added'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/tasks/<int:task_id>', methods=['PUT'])
def edit_task(task_id):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.get_json()
    task_name = data.get('task_name')
    description = data.get('description', '')
    if not task_name:
        return jsonify({'success': False, 'message': 'Task name required'}), 400
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute('UPDATE tasks SET task_name = %s, description = %s WHERE id = %s', (task_name, description, task_id))
        conn.commit()
        return jsonify({'success': True, 'message': 'Task updated'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/admin/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # Get the task name before deleting (for tables using task_name)
        cursor.execute('SELECT task_name FROM tasks WHERE id = %s', (task_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'message': 'Task not found'}), 404
        task_name = row[0]

        # Delete from user_tasks
        cursor.execute('DELETE FROM user_tasks WHERE task_name = %s', (task_name,))
        # Delete from audio_recordings (if task_name is used)
        cursor.execute('DELETE FROM audio_recordings WHERE task_name = %s', (task_name,))
        # Add more deletes here if other tables reference task_name or task_id

        # Finally, delete from tasks table
        cursor.execute('DELETE FROM tasks WHERE id = %s', (task_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Task deleted'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/users', methods=['GET'])
def admin_users_list():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    conn = connect_db()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    cursor = conn.cursor(dictionary=True)
    try:
        # Join users and demographics
        cursor.execute('''
            SELECT u.id, u.name, u.email, d.age, d.gender, d.dyslexia_status, d.education_level, d.native_language, u.created_at
            FROM users u
            LEFT JOIN demographics d ON u.id = d.user_id
        ''')
        users = cursor.fetchall()
        # Define the set of all tasks
        all_tasks = ['Reading Aloud Task 1', 'Typing Task', 'Reading Comprehension', 'Mathematical Comprehension','Writing Task','Aptitude Test']
        total_tasks = len(all_tasks)
        for user in users:
            # Get completed tasks for this user (only those in all_tasks)
            cursor.execute("""
                SELECT task_name FROM user_tasks WHERE user_id = %s AND status = 'Completed'
            """, (user['id'],))
            completed_tasks = [row['task_name'] for row in cursor.fetchall() if row['task_name'] in all_tasks]
            user['progress'] = int((len(completed_tasks) / total_tasks) * 100) if total_tasks else 0
            user['dyslexia_status'] = user.get('dyslexia_status', 'N/A')
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        print(f"Error fetching users: {e}")
        return jsonify({'success': False, 'message': 'Error fetching users'}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
def admin_user_detail(user_id):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    conn = connect_db()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    cursor = conn.cursor(dictionary=True)
    try:
        # Join users and demographics for full details
        cursor.execute('''
            SELECT u.id, u.name, u.email, d.age, d.gender, d.dyslexia_status, d.education_level, d.native_language, u.created_at
            FROM users u
            LEFT JOIN demographics d ON u.id = d.user_id
            WHERE u.id = %s
        ''', (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        # Fetch user task progress
        cursor.execute('SELECT task_name, status FROM user_tasks WHERE user_id = %s', (user_id,))
        tasks = cursor.fetchall()
        user['tasks'] = tasks
        return jsonify({'success': True, 'user': user})
    except Exception as e:
        print(f"Error fetching user detail: {e}")
        return jsonify({'success': False, 'message': 'Error fetching user detail'}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/users-grouped', methods=['GET'])
def admin_users_grouped():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    conn = connect_db()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    cursor = conn.cursor(dictionary=True)
    try:
        # Parents
        cursor.execute('''
            SELECT u.id, u.name, u.email, u.created_at
            FROM users u
            LEFT JOIN demographics d ON u.id = d.user_id
            WHERE u.user_type = 'parent'
        ''')
        parents = cursor.fetchall()

        # Children
        cursor.execute('''
            SELECT u.id, u.name, u.email, d.age, d.gender, d.dyslexia_status, d.education_level, d.native_language, u.created_at
            FROM users u
            LEFT JOIN demographics d ON u.id = d.user_id
            WHERE u.user_type = 'child'
        ''')
        children = cursor.fetchall()

        # Schools
        cursor.execute('''
            SELECT s.id, s.name, s.email, s.address, s.phone, s.created_at
            FROM schools s
            ORDER BY s.created_at DESC
        ''')
        schools = cursor.fetchall()

        # Compute progress for parents and children
        all_tasks = ['Reading Aloud Task 1', 'Typing Task', 'Reading Comprehension', 'Mathematical Comprehension','Writing Task','Aptitude Test']
        total_tasks = len(all_tasks)
        def attach_progress(users):
            for u in users:
                cursor.execute("""
                    SELECT task_name FROM user_tasks WHERE user_id = %s AND status = 'Completed'
                """, (u['id'],))
                completed = [row['task_name'] for row in cursor.fetchall() if row['task_name'] in all_tasks]
                u['progress'] = int((len(completed) / total_tasks) * 100) if total_tasks else 0
                u['dyslexia_status'] = u.get('dyslexia_status', 'N/A')
        attach_progress(parents)
        attach_progress(children)

        # Attach counts to schools
        for s in schools:
            # number of parents in this school
            cursor.execute("SELECT COUNT(*) AS c FROM users WHERE user_type = 'parent' AND school_id = %s", (s['id'],))
            s['num_parents'] = cursor.fetchone()['c']
            # number of children linked to those parents
            cursor.execute('''
                SELECT COUNT(DISTINCT pc.child_id) AS c
                FROM parent_children pc
                JOIN users p ON p.id = pc.parent_id
                WHERE p.school_id = %s
            ''', (s['id'],))
            s['num_children'] = cursor.fetchone()['c']

        # return jsonify({'success': True, 'parents': parents, 'children': children, 'schools': schools})
        combined_users=(parents or []) + (children or [])
        return jsonify({'success':True, 'parents':parents, 'children': children, 'schools': schools, 'users': combined_users})
    except Exception as e:
        print(f"Error fetching grouped users: {e}")
        return jsonify({'success': False, 'message': 'Error fetching grouped users'}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/users/<int:user_id>/export', methods=['GET'])
def admin_user_export(user_id):
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    conn = connect_db()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT id, name, email, age, gender, dyslexia_status, education_level, native_language, created_at, status
            FROM users WHERE id = %s
        """, (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        cursor.execute("""
            SELECT task_name, status FROM user_tasks WHERE user_id = %s
        """, (user_id,))
        tasks = cursor.fetchall()
        user['tasks'] = tasks
        import json
        from flask import make_response
        response = make_response(json.dumps(user, indent=2))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename=user_{user_id}_data.json'
        return response
    except Exception as e:
        print(f"Error exporting user data: {e}")
        return jsonify({'success': False, 'message': 'Error exporting user data'}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/dashboard-stats', methods=['GET'])
def admin_dashboard_stats():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    conn = connect_db()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed'}), 500
    cursor = conn.cursor(dictionary=True)
    stats = {}  # <-- Initialize stats dictionary
    try:
        # Total participants
        cursor.execute("SELECT COUNT(*) as total FROM users WHERE user_type = 'child'")
        total_participants = cursor.fetchone()['total']
        stats['totalParticipants'] = total_participants

        # Average completion rate (average progress across all users)
        cursor.execute('SELECT id FROM users')
        user_ids = [row['id'] for row in cursor.fetchall()]
        total_progress = 0
        for user_id in user_ids:
            cursor.execute('SELECT COUNT(*) as total FROM user_tasks WHERE user_id = %s', (user_id,))
            total_tasks = cursor.fetchone()['total']
            cursor.execute("SELECT COUNT(*) as completed FROM user_tasks WHERE user_id = %s AND status = 'Completed'", (user_id,))
            completed_tasks = cursor.fetchone()['completed']
            progress = int((completed_tasks / total_tasks) * 100) if total_tasks else 0
            total_progress += progress
        avg_completion = int(total_progress / len(user_ids)) if user_ids else 0
        stats['completionRate'] = avg_completion

        # Active studies (number of tasks)
        cursor.execute('SELECT COUNT(*) AS numTasks FROM tasks')
        stats['activeStudies'] = cursor.fetchone()['numTasks']

        # Data quality placeholder (set to 85 for now)
        stats['dataQuality'] = 85

        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        print(f"Error fetching dashboard stats: {e}")
        return jsonify({'success': False, 'message': 'Error fetching dashboard stats'}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/admin/recent-activity', methods=['GET'])
def admin_recent_activity():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        cur = conn.cursor(dictionary=True)

        activities = []

        # New registrations (users)
        try:
            cur.execute("""
                SELECT id, name, created_at
                FROM users
                ORDER BY created_at DESC
                LIMIT 20
            """)
            for row in cur.fetchall():
                activities.append({
                    'type': 'registration',
                    'description': f"New user registered: {row.get('name') or 'User'} (ID {row['id']})",
                    'timestamp': str(row['created_at'])
                })
        except Exception:
            pass

        # Task updates
        try:
            cur.execute("""
                SELECT user_id, task_name, status, updated_at
                FROM user_tasks
                ORDER BY updated_at DESC
                LIMIT 50
            """)
            for row in cur.fetchall():
                typ = 'completion' if row['status'] == 'Completed' else 'progress'
                activities.append({
                    'type': typ,
                    'description': f"Task {row['task_name']} → {row['status']} (User {row['user_id']})",
                    'timestamp': str(row['updated_at'])
                })
        except Exception:
            pass

        # Audio uploads
        try:
            cur.execute("""
                SELECT user_id, filename, uploaded_at
                FROM audio_recordings
                ORDER BY uploaded_at DESC
                LIMIT 50
            """)
            for row in cur.fetchall():
                activities.append({
                    'type': 'audio',
                    'description': f"Audio uploaded by User {row['user_id']}: {row['filename']}",
                    'timestamp': str(row['uploaded_at'])
                })
        except Exception:
            pass

        # Writing uploads
        try:
            cur.execute("""
                SELECT user_id, filename, uploaded_at
                FROM writing_samples
                ORDER BY uploaded_at DESC
                LIMIT 50
            """)
            for row in cur.fetchall():
                activities.append({
                    'type': 'writing',
                    'description': f"Writing sample uploaded by User {row['user_id']}: {row['filename']}",
                    'timestamp': str(row['uploaded_at'])
                })
        except Exception:
            pass

        # Sort and limit
        activities.sort(key=lambda a: a['timestamp'], reverse=True)
        activities = activities[:50]

        cur.close()
        conn.close()
        return jsonify({'success': True, 'activities': activities})
    except Exception as e:
        print(f"Recent activity error: {e}")
        return jsonify({'success': False, 'message': 'Failed to load recent activity'}), 500

@app.route('/api/admin/task-stats', methods=['GET'])
def admin_task_stats():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        cur = conn.cursor(dictionary=True)

        # Get all tasks
        cur.execute("SELECT task_name FROM tasks ORDER BY id")
        task_names = [r['task_name'] for r in cur.fetchall()]
        if not task_names:
            return jsonify({'success': True, 'labels': [], 'data': []})

        # Number of participants (children)
        cur.execute("SELECT COUNT(*) AS c FROM users WHERE user_type='child'")
        total_participants = cur.fetchone()['c'] or 0

        labels = []
        data = []
        for t in task_names:
            labels.append(t)
            if total_participants == 0:
                data.append(0)
                continue
            cur.execute("""
                SELECT COUNT(*) AS c FROM user_tasks
                WHERE task_name=%s AND status='Completed'
            """, (t,))
            completed = cur.fetchone()['c'] or 0
            pct = int(round((completed / total_participants) * 100))
            data.append(pct)

        cur.close()
        conn.close()
        return jsonify({'success': True, 'labels': labels, 'data': data})
    except Exception as e:
        print(f"Task stats error: {e}")
        return jsonify({'success': False, 'message': 'Failed to load task stats'}), 500

@app.route('/api/admin/get-child-tasks', methods=['GET'])
def admin_get_child_tasks():
    """Get all tasks for a specific child user"""
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': 'User ID required'}), 400
    
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        
        # Get all available tasks
        cursor.execute("SELECT task_name FROM tasks ORDER BY id")
        all_tasks = [row['task_name'] for row in cursor.fetchall()]
        
        # Get current task statuses for this user
        cursor.execute("SELECT task_name, status FROM user_tasks WHERE user_id = %s", (user_id,))
        user_tasks = {row['task_name']: row['status'] for row in cursor.fetchall()}
        
        # Create task list with current status or 'Not Started' as default
        tasks = []
        for task_name in all_tasks:
            status = user_tasks.get(task_name, 'Not Started')
            tasks.append({'task_name': task_name, 'status': status})
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'tasks': tasks})
        
    except Exception as e:
        print(f"Error getting child tasks: {e}")
        return jsonify({'success': False, 'message': f'Failed to get tasks: {str(e)}'}), 500




@app.route('/api/get-writing-progress', methods=['GET'])
def get_writing_progress():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'User not logged in'}), 401
    task_id = request.args.get('task_id')
    if not task_id:
        return jsonify({'success': False, 'message': 'No task_id provided'}), 400
    try:
        conn = connect_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT filename, status FROM writing_samples
            WHERE user_id = %s AND task_id = %s
            ORDER BY uploaded_at DESC LIMIT 1
        """, (session['user_id'], task_id))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        if result:
            return jsonify({'success': True, 'progress': result})
        else:
            return jsonify({'success': True, 'progress': None})
    except Exception as e:
        print(f"Get writing progress error: {e}")
        return jsonify({'success': False, 'message': 'Failed to get writing progress'}), 500


@app.route('/api/admin/export-dataset', methods=['POST'])
def admin_export_dataset():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    payload = request.get_json() or {}
    export_format = (payload.get('format') or 'json').lower()
    anonymization = (payload.get('anonymization') or 'full').lower()
    start_date_str = payload.get('startDate')
    end_date_str = payload.get('endDate')
    include = payload.get('includeData') or {}

    def parse_date(d):
        try:
            if not d:
                return None
            return datetime.strptime(d, '%Y-%m-%d')
        except Exception:
            return None

    from datetime import timedelta
    start_dt = parse_date(start_date_str)
    end_dt = parse_date(end_date_str)
    if end_dt:
        end_dt = end_dt + timedelta(days=1)

    def mask_name(name: str) -> str:
        if not name:
            return name
        parts = name.split()
        masked = []
        for p in parts:
            if len(p) <= 2:
                masked.append('*' * len(p))
            else:
                masked.append(p[0] + ('*' * (len(p) - 2)) + p[-1])
        return ' '.join(masked)

    def mask_email(email: str) -> str:
        if not email or '@' not in email:
            return email
        user, domain = email.split('@', 1)
        if len(user) <= 2:
            user_masked = '*' * len(user)
        else:
            user_masked = user[0] + ('*' * (len(user) - 2)) + user[-1]
        return f"{user_masked}@{domain}"

    def anonymize_row(row: dict) -> dict:
        if anonymization == 'none':
            return row
        r = dict(row)
        if anonymization == 'partial':
            if 'name' in r:
                r['name'] = mask_name(r['name'])
            if 'email' in r:
                r['email'] = mask_email(r['email'])
        else:
            for key in ['name', 'email', 'ip_address', 'user_agent', 'address', 'phone']:
                if key in r:
                    r[key] = None
        return r

    try:
        conn = connect_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        cur = conn.cursor(dictionary=True)

        data = {}

        # Demographics
        if include.get('demographics', False):
            where = []
            params = []
            if start_dt:
                where.append('d.created_at >= %s')
                params.append(start_dt)
            if end_dt:
                where.append('d.created_at < %s')
                params.append(end_dt)
            where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
            cur.execute(f'''
                SELECT u.id AS user_id, u.name, u.email, u.user_type, d.age, d.gender, d.native_language, d.education_level, d.dyslexia_status, d.created_at
                FROM users u
                LEFT JOIN demographics d ON u.id=d.user_id
                {where_sql}
            ''', tuple(params))
            rows = [anonymize_row(r) for r in cur.fetchall()]
            data['demographics'] = rows

        # Audio
        if include.get('audio', False):
            where = []
            params = []
            if start_dt:
                where.append('uploaded_at >= %s')
                params.append(start_dt)
            if end_dt:
                where.append('uploaded_at < %s')
                params.append(end_dt)
            where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
            cur.execute(f'''SELECT user_id, filename, task_name, uploaded_at FROM audio_recordings {where_sql} ORDER BY uploaded_at DESC''', tuple(params))
            rows = cur.fetchall()
            data['audio'] = rows

        # Typing
        if include.get('typing', False):
            where = []
            params = []
            if start_dt:
                where.append('updated_at >= %s')
                params.append(start_dt)
            if end_dt:
                where.append('updated_at < %s')
                params.append(end_dt)
            where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
            cur.execute(f'''SELECT user_id, text, keystrokes, timer, updated_at FROM typing_progress {where_sql}''', tuple(params))
            rows = cur.fetchall()
            data['typing'] = rows

        # Comprehension
        if include.get('comprehension', False):
            where = []
            params = []
            if start_dt:
                where.append('updated_at >= %s')
                params.append(start_dt)
            if end_dt:
                where.append('updated_at < %s')
                params.append(end_dt)
            where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
            cur.execute(f'''SELECT user_id, q1, q2, q3, status, updated_at FROM comprehension_progress {where_sql}''', tuple(params))
            rows = cur.fetchall()
            data['comprehension'] = rows

        # Progress
        if include.get('progress', False):
            where = []
            params = []
            if start_dt:
                where.append('updated_at >= %s')
                params.append(start_dt)
            if end_dt:
                where.append('updated_at < %s')
                params.append(end_dt)
            where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
            cur.execute(f'''SELECT user_id, task_name, status, updated_at FROM user_tasks {where_sql}''', tuple(params))
            rows = cur.fetchall()
            data['progress'] = rows

        cur.close()
        conn.close()

        if export_format == 'json':
            from flask import send_file
            out = BytesIO(json.dumps(data, ensure_ascii=False, indent=2, default=str).encode('utf-8'))
            out.seek(0)
            return send_file(out, mimetype='application/json', as_attachment=True, download_name=f"dyslexia_study_data_{datetime.utcnow().date().isoformat()}.json")

        if export_format == 'csv':
            import csv
            text_buf = StringIO()
            writer = csv.writer(text_buf)
            for section_name, rows in data.items():
                writer.writerow([f'== {section_name.upper()} =='])
                if rows:
                    header = list(rows[0].keys())
                    writer.writerow(header)
                    for r in rows:
                        writer.writerow([str(r.get(k, '')) for k in header])
                else:
                    writer.writerow(['(no rows)'])
                writer.writerow([])
            out = BytesIO(text_buf.getvalue().encode('utf-8'))
            from flask import send_file
            return send_file(out, mimetype='text/csv', as_attachment=True, download_name=f"dyslexia_study_data_{datetime.utcnow().date().isoformat()}.csv")

        if export_format in ('excel', 'xlsx'):
            try:
                from openpyxl import Workbook
            except Exception:
                # fallback to csv
                import csv
                text_buf = StringIO()
                writer = csv.writer(text_buf)
                for section_name, rows in data.items():
                    writer.writerow([f'== {section_name.upper()} =='])
                    if rows:
                        header = list(rows[0].keys())
                        writer.writerow(header)
                        for r in rows:
                            writer.writerow([str(r.get(k, '')) for k in header])
                    else:
                        writer.writerow(['(no rows)'])
                    writer.writerow([])
                out = BytesIO(text_buf.getvalue().encode('utf-8'))
                from flask import send_file
                return send_file(out, mimetype='text/csv', as_attachment=True, download_name=f"dyslexia_study_data_{datetime.utcnow().date().isoformat()}.csv")

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Summary'
            row_idx = 1
            ws.cell(row=row_idx, column=1, value='Datasets included:')
            row_idx += 2
            for section_name in data.keys():
                ws.cell(row=row_idx, column=1, value=section_name)
                row_idx += 1
            for section_name, rows in data.items():
                sheet = wb.create_sheet(title=section_name[:31])
                if rows:
                    header = list(rows[0].keys())
                    for ci, col_name in enumerate(header, start=1):
                        sheet.cell(row=1, column=ci, value=col_name)
                    for ri, r in enumerate(rows, start=2):
                        for ci, col_name in enumerate(header, start=1):
                            sheet.cell(row=ri, column=ci, value=r.get(col_name))
                else:
                    sheet.cell(row=1, column=1, value='(no rows)')
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            from flask import send_file
            return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"dyslexia_study_data_{datetime.utcnow().date().isoformat()}.xlsx")

        return jsonify({'success': False, 'message': f'Unsupported export format: {export_format}'}), 400

    except Exception as e:
        print(f'Admin export error: {e}')
        return jsonify({'success': False, 'message': 'Failed to export dataset'}), 500
@app.route('/api/admin/set-user-task-status', methods=['POST'])
def admin_set_user_task_status():
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.get_json()
    user_id = data.get('user_id')
    task_name = data.get('task_name')
    status = data.get('status')
    if not user_id or not task_name or not status:
        return jsonify({'success': False, 'message': 'Missing fields'}), 400
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO user_tasks (user_id, task_name, status)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE status = VALUES(status), updated_at = CURRENT_TIMESTAMP
        ''', (user_id, task_name, status))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/stats')
def stats():
    return render_template('stats.html')

@app.route('/admin/logout')
def admin_logout_page():
    session.pop('is_admin', None)
    flash('Admin logged out successfully', 'success')
    return redirect(url_for('landing'))

@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    session.pop('is_admin', None)
    return '', 204  # No content, JS will handle redirect

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)